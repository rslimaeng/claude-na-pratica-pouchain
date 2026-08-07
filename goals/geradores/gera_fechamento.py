# -*- coding: utf-8 -*-
"""
Planilha do EXERCICIO da aula 1.2 · Financeiro da Grafica Aurora.
Duas abas, junho e julho de 2026, 118 linhas de conta somadas.

O tamanho e o argumento da aula. Em 15 linhas o aluno acha as causas passando o
olho e a tese cai. Em 118, com conta renomeada e rateio no meio, achar na mao
deixa de ser demorado e passa a ser inviavel.

Tres armadilhas, e as tres SO existem por causa do volume:

  1. CONTA RENOMEADA  Tres contas mudaram de nome de um mes para o outro. Quem
                      nao percebe reporta "conta sumiu" e "conta nova apareceu"
                      como duas das maiores causas, e as duas sao a mesma coisa.
  2. RATEIO           Os rateios variam MAIS que qualquer causa real. Quem nao
                      exclui devolve tres causas que nao sao decisao de ninguem.
  3. CENTRO EM BRANCO Parte das linhas esta sem centro de custo. Somar por
                      centro sem tratar isso perde dinheiro no caminho.

Sujeira estrutural: cabecalho mesclado, valor ora numero ora texto, coluna com
espaco no nome, conta com espaco duplo, linha em branco no meio de cada aba.
"""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

random.seed(2481)

# (conta em junho, conta em julho, centro de custo, valor junho, valor julho)
# quando os dois nomes diferem, e a MESMA conta: armadilha 1
CONTAS = [
    # ── receita ────────────────────────────────────────────────────────────
    ("Receita de impressão offset", None, "Comercial", 1_140_000.00, 1_105_000.00),
    ("Receita de impressão digital", None, "Comercial", 146_800.00, 176_000.00),
    ("Receita de impressão flexo", None, "Comercial", 256_300.00, 246_000.00),
    ("Receita de acabamento", None, "Acabamento", 116_900.00, 109_200.00),
    ("Receita de pré-impressão", None, "Pré-impressão", 51_100.00, 49_400.00),
    ("Receita de serviços gráficos diversos", "Receita de serviços diversos", "Comercial", 187_600.00, 96_400.00),
    # ── deduções ───────────────────────────────────────────────────────────
    ("ICMS sobre vendas", None, None, -186_000.00, -181_000.00),
    ("PIS sobre vendas", None, None, -11_300.00, -10_700.00),
    ("COFINS sobre vendas", None, None, -52_000.00, -49_100.00),
    ("ISS sobre serviços", None, None, -9_200.00, -8_600.00),
    ("Devoluções e abatimentos", None, "Comercial", -10_200.00, -26_900.00),
    # ── matéria-prima ──────────────────────────────────────────────────────
    ("Papel couché", None, "Produção", -198_400.00, -271_900.00),
    ("Papel offset", None, "Produção", -57_900.00, -61_300.00),
    ("Cartão duplex e triplex", None, "Produção", -75_300.00, -71_500.00),
    ("Papel kraft", None, "Produção", -11_500.00, -12_100.00),
    ("Tinta escala", None, "Produção", -35_000.00, -37_100.00),
    ("Tinta especial Pantone", None, "Produção", -13_100.00, -16_500.00),
    ("Verniz e revestimento", None, "Acabamento", -19_300.00, -19_900.00),
    ("Chapa CTP", None, "Pré-impressão", -25_800.00, -27_100.00),
    ("Filme de laminação", None, "Acabamento", -17_800.00, -18_500.00),
    ("Cola, wire-o e fita", None, "Acabamento", -10_100.00, -10_700.00),
    ("Solvente e material de limpeza", None, "Produção", -5_800.00, -6_100.00),
    # ── mão de obra ────────────────────────────────────────────────────────
    ("Salários produção", None, "Produção", -96_000.00, -96_800.00),
    ("Salários acabamento", None, "Acabamento", -32_400.00, -32_800.00),
    ("Salários pré-impressão", None, "Pré-impressão", -20_200.00, -20_400.00),
    ("Salários comercial", None, "Comercial", -37_100.00, -37_900.00),
    ("Salários administrativo", None, "Administrativo", -43_800.00, -44_300.00),
    ("Horas extras de produção", "Horas extras fabris", "Produção", -18_400.00, -71_300.00),
    ("Encargos sociais", None, None, -69_000.00, -77_400.00),
    ("Benefícios e vale-transporte", None, None, -21_700.00, -22_500.00),
    ("Provisão de férias e 13º", None, None, -25_100.00, -27_100.00),
    ("Rescisões e verbas", None, "Administrativo", -3_200.00, -7_700.00),
    # ── ocupação e manutenção ──────────────────────────────────────────────
    ("Energia elétrica", "Energia", "Produção", -50_600.00, -55_400.00),
    ("Água e esgoto", None, None, -3_000.00, -3_200.00),
    ("Aluguel do parque fabril", None, "Produção", -68_100.00, -68_100.00),
    ("Condomínio e IPTU", None, None, -8_400.00, -8_400.00),
    ("Manutenção de máquinas", None, "Produção", -25_300.00, -29_300.00),
    ("Peças de reposição", None, "Produção", -13_800.00, -19_500.00),
    ("Manutenção predial", None, None, -5_400.00, -5_100.00),
    ("Seguro do parque fabril", None, None, -10_800.00, -10_800.00),
    # ── despesas comerciais ────────────────────────────────────────────────
    ("Comissões de vendas", None, "Comercial", -41_200.00, -37_000.00),
    ("Fretes sobre vendas", None, "Comercial", -33_800.00, -31_900.00),
    ("Marketing e feiras", None, "Comercial", -10_900.00, -8_600.00),
    ("Viagens e representação", None, "Comercial", -7_800.00, -6_700.00),
    ("Brindes e amostras", None, "Comercial", -2_800.00, -2_500.00),
    # ── despesas administrativas ───────────────────────────────────────────
    ("Software e licenças", None, "Administrativo", -18_400.00, -18_400.00),
    ("Telefonia e internet", None, "Administrativo", -5_800.00, -5_900.00),
    ("Honorários contábeis", None, "Administrativo", -9_600.00, -9_600.00),
    ("Honorários jurídicos", None, "Administrativo", -4_200.00, -18_700.00),
    ("Material de escritório", None, "Administrativo", -2_800.00, -3_100.00),
    ("Treinamento e desenvolvimento", None, "Administrativo", -6_400.00, -11_200.00),
    ("Despesas com veículos", None, "Administrativo", -9_700.00, -10_100.00),
    # ── financeiro ─────────────────────────────────────────────────────────
    ("Juros de empréstimo", None, None, -22_400.00, -23_100.00),
    ("Tarifas bancárias", None, None, -3_900.00, -4_100.00),
    ("Descontos concedidos", None, "Comercial", -7_100.00, -6_400.00),
    ("Receita de aplicação financeira", None, None, 4_800.00, 3_900.00),
    # ── rateios · armadilha 2 ──────────────────────────────────────────────
    ("Rateio administrativo", None, "Rateio", -84_600.00, -162_300.00),
    ("Rateio de estrutura fabril", None, "Rateio", -118_200.00, -41_700.00),
    ("Rateio de TI corporativo", None, "Rateio", -22_400.00, -68_900.00),
]
RENOMEADAS = [(a, b) for a, b, *_ in CONTAS if b]
RATEIO = {c[0] for c in CONTAS if c[2] == "Rateio"}
# linhas que ficam sem centro de custo no arquivo · armadilha 3
SEM_CENTRO = {i for i, c in enumerate(CONTAS) if c[2] is None}

MESES = [("Junho 2026", 3, "30/06/2026"), ("Julho 2026", 4, "31/07/2026")]


def moeda(v):
    s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"-{s}" if v < 0 else s


wb = Workbook()
wb.remove(wb.active)
n_texto = n_linhas = 0

for aba, col_valor, fecho in MESES:
    ws = wb.create_sheet(aba)
    ws.merge_cells("A1:E1")
    ws["A1"] = "GRÁFICA AURORA · FECHAMENTO GERENCIAL"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Competência {aba} · posição em {fecho} · exportado do ERP"
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    for i, c in enumerate(["Conta", "Centro de custo", " Valor ", "Natureza",
                           "Competência"], start=1):
        cel = ws.cell(row=4, column=i, value=c)
        cel.font = Font(bold=True, size=10)
        cel.fill = PatternFill("solid", fgColor="DDDDDD")

    linha = 5
    for idx, (nome_jun, nome_jul, centro, v_jun, v_jul) in enumerate(CONTAS):
        nome = nome_jun if col_valor == 3 else (nome_jul or nome_jun)
        valor = v_jun if col_valor == 3 else v_jul
        if idx % 7 == 3:                       # espaço duplo, sujeira de ERP
            nome = nome.replace(" ", "  ", 1)
        texto = (idx % 5 == 2)                 # valor gravado como texto
        if texto:
            n_texto += 1
        ws.cell(row=linha, column=1, value=nome)
        ws.cell(row=linha, column=2,
                value=None if idx in SEM_CENTRO else centro)
        ws.cell(row=linha, column=3, value=moeda(valor) if texto else valor)
        ws.cell(row=linha, column=4,
                value="Receita" if valor > 0 else "Despesa")
        ws.cell(row=linha, column=5, value=aba)
        linha += 1
        n_linhas += 1
        if idx == 26:                          # linha em branco no meio
            linha += 1

    for i, w in enumerate([38, 18, 16, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

destino = ("/Users/rafaellima/developer/4-cursos-treinamentos/treinamentos-in-company/"
           "pouchain-claude-na-pratica/site/m1/a2-pedir-para-entregar/exercicio/"
           "fechamento-dois-meses.xlsx")
wb.save(destino)

# ── os fatos que as páginas vão afirmar ─────────────────────────────────────
print(f"gravado em {destino}\n")
print(f"{'linhas de conta somadas':34} {n_linhas}")
print(f"{'contas por mês':34} {len(CONTAS)}")
print(f"{'valores gravados como texto':34} {n_texto}")
print(f"{'linhas sem centro de custo':34} {len(SEM_CENTRO) * 2}")
print(f"{'contas renomeadas entre os meses':34} {len(RENOMEADAS)}")
for a, b in RENOMEADAS:
    print(f"{'':34} \"{a}\"  ->  \"{b}\"")

res_jun = sum(c[3] for c in CONTAS)
res_jul = sum(c[4] for c in CONTAS)
print(f"\n{'resultado de junho':34} {res_jun:>14,.2f}")
print(f"{'resultado de julho':34} {res_jul:>14,.2f}")
print(f"{'variação':34} {res_jul - res_jun:>14,.2f}")

print("\n── se você INCLUIR os rateios (resposta errada) ──")
todas = sorted(CONTAS, key=lambda c: abs(c[4] - c[3]), reverse=True)
for c in todas[:3]:
    print(f"   {c[0][:38]:40} {c[4] - c[3]:>13,.2f}")

print("\n── excluindo rateio, que é o que o pedido manda (resposta certa) ──")
reais = [c for c in CONTAS if c[0] not in RATEIO]
for c in sorted(reais, key=lambda x: abs(x[4] - x[3]), reverse=True)[:3]:
    print(f"   {c[0][:38]:40} {c[4] - c[3]:>13,.2f}")

soma_rateio = sum(c[4] - c[3] for c in CONTAS if c[0] in RATEIO)
print(f"\nvariação só dos rateios: {soma_rateio:,.2f}  "
      f"(some isso e a conta do resultado não fecha com a operação)")
print("as 3 renomeadas, se lidas como contas diferentes, viram 6 causas falsas:")
for a, b in RENOMEADAS:
    va = next(c[3] for c in CONTAS if c[0] == a)
    vb = next(c[4] for c in CONTAS if c[0] == a)
    print(f"   \"{a[:34]}\" some ({va:,.2f}) e \"{b[:30]}\" nasce ({vb:,.2f})")
