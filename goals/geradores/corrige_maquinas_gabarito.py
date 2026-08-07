# -*- coding: utf-8 -*-
"""
Conserto de fato no GABARITO da aula 1.3 · minhas-regras-GABARITO.docx

O gabarito dizia "Temos 4 maquinas: Offset 1, Offset 2, Flexo 1 e Flexo 2".
A planilha do exercicio, pedidos-em-producao.xlsx, tem SEIS. As mesmas seis
aparecem no prompt situado da aula 1.1 e na saida da demonstracao da 1.3, na
mesma pagina.

Por que isso nao era erro cosmetico: o exercicio manda colar essas regras nas
Instrucoes de um Project e rodar o mesmo pedido vago na MESMA planilha. Quem
usasse o gabarito de referencia ensinaria ao Claude que Digital 1 e Offset 3
nao existem, e a aula que promete melhorar a resposta entregaria uma pior.

Este script le a lista de maquinas DA PLANILHA e escreve nela o gabarito, para
os dois nunca mais divergirem. Roda de novo sem estragar nada: se o texto ja
estiver certo, ele avisa e nao grava.
"""
import re
import openpyxl
from docx import Document

BASE = ("/Users/rafaellima/developer/4-cursos-treinamentos/treinamentos-in-company/"
        "pouchain-claude-na-pratica/site/m1/")
PLANILHA = BASE + "a1-ecossistema-e-fisica/exercicio/pedidos-em-producao.xlsx"
GABARITO = BASE + "a3-regra-que-fica/gabarito/minhas-regras-GABARITO.docx"

# ── a verdade sai da planilha, não do texto ─────────────────────────────────
ws = openpyxl.load_workbook(PLANILHA, data_only=True).active
cab = [c for c in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
i_maq = [i for i, c in enumerate(cab) if c and "quina" in str(c)][0]
achadas = {r[i_maq] for r in ws.iter_rows(min_row=5, values_only=True)
           if r and r[0] and r[i_maq]}
# ordem de chão de fábrica, a mesma do prompt situado da 1.1, não a alfabética
ORDEM = ["Offset", "Flexo", "Digital"]
maquinas = sorted(achadas, key=lambda m: (ORDEM.index(m.split()[0]), m))
NUM = {1: "uma", 2: "duas", 3: "três", 4: "quatro", 5: "cinco", 6: "seis"}
PLURAL = {"Offset": "offset", "Flexo": "flexográficas", "Digital": "digitais"}
SING = {"Offset": "offset", "Flexo": "flexográfica", "Digital": "digital"}

familia = {}
for m in maquinas:
    familia.setdefault(m.split()[0], []).append(m)
partes = [f"{NUM[len(v)]} {PLURAL[k] if len(v) > 1 else SING[k]}"
          for k, v in familia.items()]
frase_programo = ("Programo a produção de " + ", ".join(partes[:-1]) +
                  " e " + partes[-1] + ".")
frase_temos = (f"Temos {NUM[len(maquinas)]} máquinas: " +
               ", ".join(maquinas[:-1]) + " e " + maquinas[-1] + ".")
frase_tipos = ("Offset é para tiragem alta e prazo mais longo. Flexo é para "
               "rótulo e prazo curto. Digital é tiragem pequena e prazo curto.")

TROCAS = [
    (re.compile(r"^Programo a produção de .*\.$"), frase_programo),
    (re.compile(r"^Temos \w+ máquinas: .*\.$"), frase_temos),
    (re.compile(r"^Offset é para tiragem alta.*prazo curto\.$"), frase_tipos),
]

doc = Document(GABARITO)
feitas = []
for par in doc.paragraphs:
    for padrao, novo in TROCAS:
        if padrao.match(par.text.strip()) and par.text.strip() != novo:
            if len(par.runs) != 1:
                raise SystemExit(f"PREMISSA FALHOU: {len(par.runs)} runs em "
                                 f"'{par.text[:50]}'. Conserte à mão.")
            feitas.append((par.text.strip(), novo))
            par.runs[0].text = novo

print(f"máquinas na planilha: {len(maquinas)}")
for m in maquinas:
    print(f"   {m}")
print()
if not feitas:
    print("gabarito já está certo, nada gravado.")
else:
    doc.save(GABARITO)
    print(f"gravado em {GABARITO}\n")
    for velho, novo in feitas:
        print(f"  DE:    {velho}")
        print(f"  PARA:  {novo}\n")
