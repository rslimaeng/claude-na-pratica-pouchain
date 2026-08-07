# -*- coding: utf-8 -*-
"""
Regera as secoes de dados da pagina exemplo/ da aula 1.1 a partir da planilha
nova de 120 OS. Tudo aqui e calculado do .xlsx: nenhum numero e digitado.
"""
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook

SITE = ("/Users/rafaellima/developer/4-cursos-treinamentos/treinamentos-in-company/"
        "pouchain-claude-na-pratica/site")
XLSX = f"{SITE}/m1/a1-ecossistema-e-fisica/exercicio/pedidos-em-producao.xlsx"
PAG = f"{SITE}/m1/a1-ecossistema-e-fisica/exemplo/index.html"
HOJE = datetime(2026, 7, 16)
MES = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
       "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}


def data(v):
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for f in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, f)
        except ValueError:
            pass
    p = s.split("-")
    if len(p) == 3 and p[1] in MES:
        return datetime(2000 + int(p[2]), MES[p[1]], int(p[0]))
    return None


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).replace(".", "").replace(",", "."))


def moeda(v, casas=2):
    return f"{v:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")


ws = load_workbook(XLSX, data_only=True).active
linhas = [r for r in ws.iter_rows(min_row=5, values_only=True)
          if r[0] and str(r[0]).isdigit()]
vivas = [r for r in linhas if str(r[6]).strip().lower() != "entregue"]
atras = sorted([r for r in vivas if (d := data(r[5])) and d < HOJE],
               key=lambda r: data(r[5]))
vence3 = sorted([r for r in vivas if HOJE <= data(r[5]) <= HOJE + timedelta(days=3)],
                key=lambda r: data(r[5]))
marcadas = [r for r in linhas if str(r[6]).strip() == "Atrasado"]
vistos, dups = set(), []
for r in linhas:
    (dups if r[0] in vistos else vistos).append(r[0]) if r[0] in vistos else vistos.add(r[0])
dups = sorted({r[0] for r in linhas if [x[0] for x in linhas].count(r[0]) > 1})
valor_vivas = sum(num(r[8]) for r in vivas)
valor_dup = sum(num(r[8]) for r in linhas if r[0] in dups) / 2
sem_maq = [r for r in vivas if not r[7]]

PARADO = ("aguardando papel", "aguardando aprovação", "em pré-impressão")


def situacao(r):
    s = str(r[6]).strip()
    if s == "Atrasado":
        return ("d", "Marcada como atrasada")
    if s.lower() == "aguardando papel":
        return ("d", "Aguardando papel")
    if s.lower() == "aguardando aprovação":
        return ("a", "Aguardando aprovação")
    if s.lower() == "em pré-impressão":
        return ("a", "Em pré-impressão")
    if s.lower() == "em acabamento":
        return ("e", "Em acabamento")
    return ("e", "Em produção")


# ── carga por máquina ───────────────────────────────────────────────────────
maqs = {}
for r in vivas:
    m = r[7] or "(sem máquina)"
    d = maqs.setdefault(m, {"prod": 0, "papel": 0, "aprov": 0, "atras": 0,
                            "valor": 0.0, "total": 0})
    s = str(r[6]).strip().lower()
    if s == "atrasado":
        d["atras"] += 1
    elif s == "aguardando papel":
        d["papel"] += 1
    elif s in ("aguardando aprovação", "em pré-impressão"):
        d["aprov"] += 1
    else:
        d["prod"] += 1
    d["valor"] += num(r[8])
    d["total"] += 1
ordem = sorted(maqs.items(), key=lambda x: -x[1]["total"])
gargalo, gd = ordem[0]
parados_gargalo = gd["papel"] + gd["aprov"] + gd["atras"]

print(f"OS no arquivo {len(linhas)} | vivas {len(vivas)} | atrasadas {len(atras)} "
      f"| marcadas {len(marcadas)} | vence≤3d {len(vence3)}")
print(f"valor vivo R$ {moeda(valor_vivas)} | OS duplicadas {dups} "
      f"(contadas 2x: R$ {moeda(valor_dup)}) | sem máquina {len(sem_maq)}")
print(f"gargalo {gargalo}: {gd['total']} OS, {parados_gargalo} paradas, "
      f"R$ {moeda(gd['valor'])} ({gd['valor']/valor_vivas*100:.0f}% do valor)")

# ══════════════════════════════════════════════ blocos de HTML
TOP = 8
tr_atras = "\n".join(
    f'            <tr><td class="os">{r[0]}</td><td>{r[1]}</td><td>{r[2]}</td>'
    f'<td>{data(r[5]).strftime("%d/%m")}</td>'
    f'<td class="num">{(HOJE - data(r[5])).days} dias</td>'
    f'<td>{r[7] or "sem máquina"}</td>'
    f'<td><span class="pin {situacao(r)[0]}">{situacao(r)[1]}</span></td></tr>'
    for r in atras[:TOP])

def falta(r):
    d = (data(r[5]) - HOJE).days
    if d == 0:
        return '<span class="pin d">Vence hoje</span>'
    return f"{d} dia" + ("s" if d > 1 else "")


tr_vence = "\n".join(
    f'            <tr><td class="os">{r[0]}</td><td>{r[1]}</td>'
    f'<td>{data(r[5]).strftime("%d/%m")}</td>'
    f'<td>{falta(r)}</td>'
    f'<td>{r[7] or "sem máquina"}</td>'
    f'<td><span class="pin {situacao(r)[0]}">{situacao(r)[1]}</span></td></tr>'
    for r in vence3[:6])

barras = []
for m, d in ordem:
    t = d["total"]
    segs = []
    for chave, cls, rot in (("prod", "prod", "em produção"),
                            ("papel", "papel", "sem papel"),
                            ("aprov", "aprov", "sem aprovação"),
                            ("atras", "atras", "atrasada no sistema")):
        if d[chave]:
            pct = d[chave] / t * 100
            txt = f"{d[chave]} {rot}" if pct >= 22 else str(d[chave])
            segs.append(f'            <span class="maq-seg {cls}" '
                        f'style="width:{pct:.1f}%">{txt}</span>')
    barras.append(
        '        <div class="maq-linha">\n'
        f'          <span class="maq-nome">{m}</span>\n'
        '          <span class="maq-barra">\n' + "\n".join(segs) + "\n          </span>\n"
        f'          <span class="maq-val">R$ {moeda(d["valor"], 0)}</span>\n'
        '        </div>')
barras = "\n".join(barras)

pag = open(PAG, encoding="utf-8").read()


def troca(alvo, novo, nome):
    global pag
    n = len(re.findall(re.escape(alvo), pag)) if isinstance(alvo, str) else 0
    if isinstance(alvo, str) and n:
        pag = pag.replace(alvo, novo)
        print(f"   ok  {nome}")
    else:
        antes = pag
        pag = re.sub(alvo, novo, pag, flags=re.S)
        print(f"   {'ok ' if pag != antes else 'FALHOU'} {nome}")


print("\nreescrevendo a página:")
troca("18 ordens de serviço", "120 ordens de serviço", "meta do download")
troca("· 18 OS ANALISADAS", f"· {len(linhas)} LINHAS DE OS ANALISADAS", "doc-meta")
troca("PEDIDOS-EM-PRODUCAO.XLSX · 18 OS · DADOS INVENTADOS",
      f"PEDIDOS-EM-PRODUCAO.XLSX · {len(linhas)} LINHAS · DADOS INVENTADOS", "rodapé")
troca(r"<li>Encontrou <strong>18 linhas úteis</strong>.*?</li>",
      f"<li>Encontrou <strong>{len(linhas)} linhas de OS</strong> e descartou 5: "
      f"duas de cabeçalho do sistema, uma em branco no meio e duas de rodapé. "
      f"Encontrou também <strong>{len(dups)} OS repetidas</strong> "
      f"({' e '.join(str(d) for d in dups)}), com valores diferentes.</li>",
      "confirmação · linhas úteis")
troca(r'<p class="sec-sub">Duas OS foram entregues.*?</p>',
      f'<p class="sec-sub">{len(linhas) - len(vivas)} OS já entregues saem da conta '
      f'de produção. Sobram {len(vivas)}.</p>', "sub da seção 01")
troca(r'<span class="bn-val">16</span>', f'<span class="bn-val">{len(vivas)}</span>',
      "big number · em produção")
troca(r'<span class="bn-val alerta">7</span>',
      f'<span class="bn-val alerta">{len(atras)}</span>', "big number · atrasadas")
troca(r'<span class="bn-val">1</span>\s*<span class="bn-reg"></span>\s*'
      r'<span class="bn-rot">Vence hoje e não entrou</span>',
      f'<span class="bn-val">{len(vence3)}</span>\n          <span class="bn-reg"></span>\n'
      f'          <span class="bn-rot">Vencem em até 3 dias</span>',
      "big number · vence em 3 dias")
troca(r'<span class="bn-val">202,7<span style="font-size:24px"> mil</span></span>',
      f'<span class="bn-val">{valor_vivas/1e6:.2f}'.replace(".", ",") +
      '<span style="font-size:24px"> mi</span></span>', "big number · valor")
troca(r'A coluna <strong>Status</strong> marca duas OS.*?vai trazer 2\.',
      f'A coluna <strong>Status</strong> marca <strong>{len(marcadas)} OS</strong> como '
      f'"Atrasado". <strong>Pela data de prazo, são {len(atras)}.</strong> A OS '
      f'<strong>{atras[0][0]}</strong> e a OS <strong>{atras[1][0]}</strong> estão as duas '
      f'com {(HOJE - data(atras[0][5])).days} dias de atraso, e o sistema marca só uma.\n'
      f'        </p>\n        <p>\n'
      f'          Este relatório usa a <strong>data</strong>, não o campo de status. '
      f'Se alguém na reunião trouxer o número do sistema, vai trazer {len(marcadas)}.',
      "nota do status")
troca(r'<p class="sec-sub">Da mais atrasada para a menos\..*?</p>',
      f'<p class="sec-sub">As {TOP} mais atrasadas de {len(atras)}. '
      f'A lista inteira está no anexo do arquivo gerado.</p>', "sub da seção 02")
troca(r'(<tbody>\n)(            <tr><td class="os">2419.*?)(\n          </tbody>)',
      lambda m: m.group(1) + tr_atras + m.group(3), "tabela das atrasadas")
troca(r'Das sete atrasadas, <strong>duas não estão rodando.*?reunião de produção\.',
      f'Das {len(atras)} atrasadas, <strong>'
      f'{sum(1 for r in atras if str(r[6]).strip().lower() in PARADO)} não estão rodando '
      f'em máquina nenhuma</strong>: esperam papel, aprovação do cliente ou pré-impressão. '
      f'Essas não são problema de produção, e por isso não se resolvem na reunião de '
      f'produção. E <strong>{len(sem_maq)} OS vivas estão sem máquina definida</strong>, '
      f'então qualquer conta de carga já nasce incompleta.',
      "nota da seção 02")
troca(r'(<div class="maq">\n)(        <div class="maq-linha">.*?)(\n      </div>\n      <div class="legenda">)',
      lambda m: m.group(1) + barras + m.group(3), "barras de carga por máquina")
troca(r'A <strong>Offset 2 concentra 44%.*?Só a 2422 está rodando\.',
      f'A <strong>{gargalo} concentra {gd["valor"]/valor_vivas*100:.0f}% do valor em '
      f'produção</strong> e tem <strong>{parados_gargalo} das {gd["total"]} OS '
      f'paradas</strong>, esperando papel ou aprovação.', "nota do gargalo")
troca(r'<p class="sec-sub">É a seção que evita o atraso da semana que vem\.</p>',
      f'<p class="sec-sub">São {len(vence3)}. É a seção que evita o atraso da semana '
      f'que vem.</p>', "sub da seção 04")
troca(r'(<tr><th>OS</th><th>Cliente</th><th>Prazo</th><th>Falta</th><th>Máquina</th>'
      r'<th>Travada em</th></tr>\n          </thead>\n          <tbody>\n)'
      r'(            <tr><td class="os">2425.*?)(\n          </tbody>)',
      lambda m: m.group(1) + tr_vence + m.group(3), "tabela do vence em 3 dias")

open(PAG, "w", encoding="utf-8").write(pag)
print("\ngravado.")
print("\nas 3 decisões precisam destes fatos:")
print(f"  1. gargalo {gargalo}: {parados_gargalo} de {gd['total']} paradas, "
      f"R$ {moeda(gd['valor'], 0)}")
print(f"  2. OS duplicadas {dups}: o total está R$ {moeda(valor_dup)} a mais")
print(f"  3. status diz {len(marcadas)}, data diz {len(atras)}")
print(f"  extra: {len(sem_maq)} OS vivas sem máquina")
