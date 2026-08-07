# -*- coding: utf-8 -*-
"""
Planilha do EXERCICIO da aula 1.1 · PCP da Grafica Aurora.
De 18 para 120 ordens de servico.

O motivo e a sensacao de ganho de produtividade: em 18 linhas o aluno pensa
"isso eu fazia na mao", e a segunda conversa vira luxo. Em 120, com OS
duplicada e cliente escrito de dois jeitos, fazer na mao deixa de ser
demorado e passa a ser inviavel. E ai a diferenca entre as duas conversas
para de ser argumento e vira experiencia.

Cinco armadilhas, e as tres ultimas so existem por causa do volume:
  1. STATUS MENTE   o campo Status marca poucas como "Atrasado", mas o Prazo
                    ja venceu em muito mais que nao foram entregues.
  2. TIRAGEM TEXTO  parte das tiragens esta gravada como texto ("12.000"),
                    entao somar direto ignora essas linhas.
  3. OS DUPLICADA   o export repetiu OS com valor diferente. Somar o total
                    conta duas vezes.
  4. CLIENTE DUPLO  "Distribuidora Sertao" tambem aparece como "Distrib.
                    Sertao". Agrupar por cliente reparte o mesmo cliente.
  5. MAQUINA VAZIA  OS em producao sem maquina preenchida. O gargalo por
                    maquina fica errado se ninguem reparar.

Sujeira estrutural de sempre: cabecalho mesclado, coluna com espaco no nome,
data em tres formatos, status em quatro caixas, valor ora numero ora texto,
linha em branco no meio.
"""
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

random.seed(2481)                      # determinismo: os numeros nao podem variar
HOJE = datetime(2026, 7, 16)

CLIENTES = ["Distribuidora Sertão", "Laticínios Vale Verde", "Farmácia Bem Estar",
            "Padaria Central", "Óticas Miranda", "Cosméticos Flor de Lis",
            "Bebidas Jangada", "Supermercados Praia Nova", "Confecções Tropical",
            "Frigorífico Boi Dourado", "Editora Céu Azul", "Clínica Santa Clara",
            "Construtora Horizonte", "Rede Sabor Caseiro"]

PRODUTOS = [("Rótulo adesivo", "flexo"), ("Caixa cartucho", "offset"),
            ("Bula", "offset"), ("Encarte 8p", "offset"),
            ("Sacola personalizada", "flexo"), ("Cartaz A2", "offset"),
            ("Folder 3 dobras", "offset"), ("Manual técnico", "offset"),
            ("Etiqueta térmica", "flexo"), ("Catálogo institucional", "offset"),
            ("Bloco de notas", "digital"), ("Cardápio laminado", "digital"),
            ("Calendário de mesa", "digital"), ("Embalagem flexível", "flexo")]

MAQ = {"offset": ["Offset 1", "Offset 2", "Offset 3"],
       "flexo":  ["Flexo 1", "Flexo 2"],
       "digital": ["Digital 1"]}

# caixas diferentes para o mesmo status: e assim que sai do sistema
VIVOS = ["Em produção", "em produção", "EM PRODUÇÃO", "Em acabamento",
         "Aguardando papel", "Aguardando aprovação", "Em pré-impressão"]
ENTREGUES = ["Entregue", "ENTREGUE", "entregue"]

N_OS = 118                              # + 2 duplicadas = 120 linhas
registros = []
for i in range(N_OS):
    os_num = 2418 + i
    cliente = random.choice(CLIENTES)
    produto, familia = random.choice(PRODUTOS)
    tiragem = random.choice([1500, 2500, 3500, 4000, 6000, 8000, 11000,
                             12000, 18000, 25000, 30000, 35000, 50000])
    entrada = HOJE - timedelta(days=random.randint(0, 20))
    prazo = entrada + timedelta(days=random.randint(4, 22))
    # 38% ja foi entregue; o resto continua vivo na casa
    entregue = random.random() < 0.38
    status = random.choice(ENTREGUES if entregue else VIVOS)
    # o campo Status so acusa atraso em 1 de cada 8 casos que de fato atrasaram
    if not entregue and prazo < HOJE and random.random() < 0.13:
        status = "Atrasado"
    maquina = random.choice(MAQ[familia])
    valor = round(random.uniform(3000, 42000), 2)
    registros.append(dict(os=os_num, cliente=cliente, produto=produto,
                          tiragem=tiragem, entrada=entrada, prazo=prazo,
                          status=status, maquina=maquina, valor=valor))

# ── armadilha 3 · duas OS repetidas com valor diferente ─────────────────────
dups = []
for idx in (17, 63):
    d = dict(registros[idx])
    d["valor"] = round(d["valor"] * random.uniform(1.05, 1.18), 2)
    dups.append((idx, d))
for offset, (idx, d) in enumerate(dups):
    registros.insert(idx + 1 + offset, d)

# ── armadilha 4 · o mesmo cliente escrito de dois jeitos ────────────────────
n_apelido = 0
for r in registros:
    if r["cliente"] == "Distribuidora Sertão" and random.random() < 0.45:
        r["cliente"] = "Distrib. Sertão"
        n_apelido += 1

# ── armadilha 5 · maquina em branco em OS que esta viva ─────────────────────
sem_maquina = 0
for r in registros:
    if r["status"] not in ENTREGUES and random.random() < 0.07:
        r["maquina"] = None
        sem_maquina += 1

# ── monta o arquivo ─────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "OS em producao"

ws.merge_cells("A1:I1")
ws["A1"] = "GRÁFICA AURORA · Relatório de ordens de serviço"
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:I2")
ws["A2"] = "Emitido em 16/07/2026 08:12 · módulo PCP · todas as OS abertas no período"
ws["A2"].font = Font(size=10, italic=True)
ws["A2"].alignment = Alignment(horizontal="center")

COLS = ["OS", "Cliente", "Produto", " Tiragem ", "Entrada", "Prazo",
        "Status", "Maquina", "Valor"]
for i, c in enumerate(COLS, start=1):
    cel = ws.cell(row=4, column=i, value=c)
    cel.font = Font(bold=True, size=10)
    cel.fill = PatternFill("solid", fgColor="DDDDDD")

FMT_DATA = ["%d/%m/%Y", "%Y-%m-%d", "%d-%b-%y"]
MES = {"Jan": "jan", "Feb": "fev", "Mar": "mar", "Apr": "abr", "May": "mai",
       "Jun": "jun", "Jul": "jul", "Aug": "ago", "Sep": "set", "Oct": "out",
       "Nov": "nov", "Dec": "dez"}


def escreve_data(d, i):
    f = FMT_DATA[i % 3]
    s = d.strftime(f)
    for en, pt in MES.items():
        s = s.replace(en, pt)
    return s


def moeda(v):
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


linha, n_tiragem_texto, n_valor_texto = 5, 0, 0
for i, r in enumerate(registros):
    ws.cell(row=linha, column=1, value=r["os"])
    ws.cell(row=linha, column=2, value=r["cliente"])
    ws.cell(row=linha, column=3, value=r["produto"])
    # armadilha 2 · 1 em cada 4 tiragens vai como texto com ponto de milhar
    if i % 4 == 1:
        ws.cell(row=linha, column=4, value=f"{r['tiragem']:,}".replace(",", "."))
        n_tiragem_texto += 1
    else:
        ws.cell(row=linha, column=4, value=r["tiragem"])
    ws.cell(row=linha, column=5, value=escreve_data(r["entrada"], i))
    ws.cell(row=linha, column=6, value=escreve_data(r["prazo"], i + 1))
    ws.cell(row=linha, column=7, value=r["status"])
    ws.cell(row=linha, column=8, value=r["maquina"])
    if i % 3 == 0:
        ws.cell(row=linha, column=9, value=moeda(r["valor"]))
        n_valor_texto += 1
    else:
        ws.cell(row=linha, column=9, value=r["valor"])
    linha += 1
    if i == 57:                          # linha em branco no meio do arquivo
        linha += 1

linha += 1
for nota in ["Relatório gerado automaticamente pelo módulo PCP. Não editar.",
             "Status é preenchido manualmente pelo líder de máquina no fim do turno."]:
    ws.cell(row=linha, column=1, value=nota).font = Font(italic=True, size=9)
    linha += 1

for i, w in enumerate([8, 26, 24, 11, 13, 13, 22, 12, 13], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

destino = ("/Users/rafaellima/developer/4-cursos-treinamentos/treinamentos-in-company/"
           "pouchain-claude-na-pratica/site/m1/a1-ecossistema-e-fisica/exercicio/"
           "pedidos-em-producao.xlsx")
wb.save(destino)

# ── os fatos que as páginas vão afirmar ─────────────────────────────────────
vivas = [r for r in registros if r["status"] not in ENTREGUES]
atrasadas = [r for r in vivas if r["prazo"] < HOJE]
marcadas = [r for r in registros if r["status"] == "Atrasado"]
vencem3 = [r for r in vivas if HOJE <= r["prazo"] <= HOJE + timedelta(days=3)]
carga = {}
for r in vivas:
    carga[r["maquina"]] = carga.get(r["maquina"], 0) + 1
valor_vivas = sum(r["valor"] for r in vivas)
os_repetidas = sorted({r["os"] for r in registros
                       if sum(1 for x in registros if x["os"] == r["os"]) > 1})
sertao = sum(1 for r in registros if "Sertão" in r["cliente"])

print(f"gravado em {destino}\n")
print(f"{'linhas de OS no arquivo':34} {len(registros)}")
print(f"{'OS distintas':34} {len({r['os'] for r in registros})}")
print(f"{'OS repetidas (armadilha 3)':34} {os_repetidas}")
print(f"{'entregues':34} {len(registros) - len(vivas)}")
print(f"{'em aberto (nao entregues)':34} {len(vivas)}")
print(f"{'ATRASADAS pela data':34} {len(atrasadas)}   <<<")
print(f"{'marcadas Atrasado no campo Status':34} {len(marcadas)}   <<< a mentira")
print(f"{'vencem em ate 3 dias':34} {len(vencem3)}")
print(f"{'tiragem gravada como texto':34} {n_tiragem_texto}")
print(f"{'valor gravado como texto':34} {n_valor_texto}")
print(f"{'linhas do cliente Sertao (2 grafias)':34} {sertao}  sendo {n_apelido} como 'Distrib.'")
print(f"{'OS viva sem maquina':34} {sem_maquina}")
print(f"{'valor total em aberto':34} R$ {moeda(valor_vivas)}")
print("\ncarga por maquina (so OS em aberto):")
for m, n in sorted(carga.items(), key=lambda x: (-x[1], str(x[0]))):
    print(f"   {str(m) if m else '(em branco)':16} {n}")
print("\nas 5 mais atrasadas:")
for r in sorted(atrasadas, key=lambda r: r["prazo"])[:5]:
    print(f"   OS {r['os']} · {r['cliente'][:24]:24} · venceu {r['prazo'].strftime('%d/%m')} "
          f"· {(HOJE - r['prazo']).days:>2}d · {r['maquina']}")
