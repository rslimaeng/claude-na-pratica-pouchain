# -*- coding: utf-8 -*-
"""
Planilha do EXERCICIO da aula 1.2 · RH e DP da Grafica Aurora.
Duas abas: 124 inscricoes e a descricao da vaga.

Sem nome de pessoa em lugar nenhum. As inscricoes sao codigo, do jeito que uma
triagem cega e feita de verdade, e do jeito que material de curso tem que ser.

O tamanho e o argumento da aula. Em 15 inscricoes o aluno separa na mao e a tese
cai. Em 124, com requisito ambiguo e escolaridade escrita de tres jeitos, separar
na mao deixa de ser demorado e passa a ser inviavel.

Tres armadilhas, e as tres SO existem por causa do volume:

  1. REQUISITO AMBIGUO  "Disponibilidade para trabalhar em turnos" tem duas
                        leituras: serve qualquer turno fixo, ou tem que rodar
                        entre eles? As duas leituras dao listas MUITO diferentes.
                        Quem nao pergunta aplica um criterio inventado em 124
                        pessoas de uma vez.
  2. DESEJAVEL VIRANDO  A vaga tem 4 obrigatorios e 3 desejaveis. Tratar tudo
     OBRIGATORIO        como obrigatorio esvazia a lista dos que atendem.
  3. TEMPO DE CASA      Ha inscritos com muitos anos na ultima empresa e ZERO em
                        producao grafica. O numero grande induz a inferir
                        senioridade que nao existe para esta vaga.

Sujeira estrutural: escolaridade escrita de tres jeitos, celula vazia, data em
dois formatos, coluna com espaco no nome, linha em branco no meio.
"""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

random.seed(2481)
N = 124

# escolaridade: o mesmo nivel escrito de tres jeitos e o que reprova
MEDIO = ["Ensino médio completo", "Médio completo", "2º grau completo"]
ABAIXO = ["Ensino médio incompleto", "Ensino fundamental completo"]
ACIMA = ["Ensino superior incompleto", "Ensino técnico completo",
         "Ensino superior completo"]

TURNO = ["Qualquer turno", "Manhã", "Tarde", "Noite", "A combinar"]
PESO_TURNO = [0.31, 0.24, 0.19, 0.14, 0.12]

CIDADE = ["Fortaleza", "Maracanaú", "Caucaia", "Eusébio", "Pacatuba",
          "Maranguape", "Sobral", "Juazeiro do Norte", "Quixadá"]
PESO_CIDADE = [0.46, 0.13, 0.12, 0.07, 0.06, 0.05, 0.045, 0.035, 0.03]
RMF = {"Fortaleza", "Maracanaú", "Caucaia", "Eusébio", "Pacatuba", "Maranguape"}

SETOR = ["Gráfica", "Embalagem", "Indústria têxtil", "Comércio varejista",
         "Logística", "Construção civil", "Alimentício", "Serviços"]

# ── obrigatórios da vaga ────────────────────────────────────────────────────
# 1. ensino médio completo
# 2. no mínimo 1 ano em produção gráfica
# 3. disponibilidade para trabalhar em turnos   ← AMBÍGUO, armadilha 1
# 4. residir em Fortaleza ou região metropolitana
# desejáveis: CNH B · curso técnico em artes gráficas · experiência Heidelberg

insc = []
for i in range(N):
    r = random.random()
    esc = (random.choice(MEDIO) if r < 0.62 else
           random.choice(ACIMA) if r < 0.83 else random.choice(ABAIXO))
    anos_gr = (0 if random.random() < 0.34 else
               random.choice([1, 1, 2, 2, 3, 4, 5, 6, 8, 11]))
    ult = random.choice(SETOR) if anos_gr == 0 else \
        random.choice(["Gráfica", "Gráfica", "Embalagem"])
    # armadilha 3: muito tempo de casa e nenhuma experiência na função
    anos_ult = (random.choice([9, 10, 11, 12, 14]) if anos_gr == 0 and
                random.random() < 0.22 else random.choice([1, 2, 2, 3, 4, 5, 7]))
    insc.append({
        "cod": f"INSC-{i + 1:03d}",
        "esc": esc,
        "gr": anos_gr,
        "turno": random.choices(TURNO, PESO_TURNO)[0],
        "cidade": random.choices(CIDADE, PESO_CIDADE)[0],
        "ult_setor": ult,
        "anos_ult": anos_ult,
        "cnh": random.choice(["B", "B", "AB", "A", "não possui", "não possui"]),
        "tecnico": "Sim" if random.random() < 0.21 else "Não",
        "heidelberg": "Sim" if (anos_gr >= 2 and random.random() < 0.38) else "Não",
        "data": ("14/07/2026" if i % 3 else "2026-07-14"),
    })

# algumas células vazias, do jeito que formulário devolve
for i in (19, 47, 88):
    insc[i]["cnh"] = None
for i in (33, 71):
    insc[i]["turno"] = None


def medio_ok(e):
    return e in MEDIO or e in ACIMA


def obrig(p, turno_estrito):
    """Devolve a lista de obrigatórios que a inscrição NÃO atende."""
    falta = []
    if not medio_ok(p["esc"]):
        falta.append("ensino médio")
    if p["gr"] < 1:
        falta.append("1 ano em produção gráfica")
    if turno_estrito:
        ok_t = p["turno"] == "Qualquer turno"
    else:
        ok_t = p["turno"] is not None and p["turno"] != "A combinar"
    if not ok_t:
        falta.append("disponibilidade de turno")
    if p["cidade"] not in RMF:
        falta.append("residir na RMF")
    return falta


# ── escreve o arquivo ───────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Inscrições"
ws.merge_cells("A1:J1")
ws["A1"] = "GRÁFICA AURORA · PROCESSO SELETIVO 2026-07"
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:J2")
ws["A2"] = ("Vaga: Auxiliar de Impressão Offset · inscrições anonimizadas por "
            "código · exportado do formulário em 15/07/2026 09:12")
ws["A2"].font = Font(size=10, italic=True)
ws["A2"].alignment = Alignment(horizontal="center")

COLS = ["Inscrição", "Escolaridade", "Anos em produção gráfica",
        " Disponibilidade ", "Cidade", "Setor da última empresa",
        "Anos na última empresa", "CNH", "Curso técnico artes gráficas",
        "Data da inscrição"]
for i, c in enumerate(COLS, start=1):
    cel = ws.cell(row=4, column=i, value=c)
    cel.font = Font(bold=True, size=10)
    cel.fill = PatternFill("solid", fgColor="DDDDDD")

linha = 5
for i, p in enumerate(insc):
    for col, v in enumerate([p["cod"], p["esc"], p["gr"], p["turno"],
                             p["cidade"], p["ult_setor"], p["anos_ult"],
                             p["cnh"], p["tecnico"], p["data"]], start=1):
        ws.cell(row=linha, column=col, value=v)
    linha += 1
    if i == 58:                                # linha em branco no meio
        linha += 1

for i, w in enumerate([12, 26, 22, 18, 18, 24, 20, 12, 26, 16], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── aba da descrição da vaga ────────────────────────────────────────────────
wv = wb.create_sheet("Descrição da vaga")
wv.column_dimensions["A"].width = 4
wv.column_dimensions["B"].width = 96
LINHAS_VAGA = [
    ("t", "AUXILIAR DE IMPRESSÃO OFFSET"),
    ("s", "Gráfica Aurora · Produção · vaga efetiva · escala 6x1"),
    ("", ""),
    ("h", "Requisitos obrigatórios"),
    ("i", "1. Ensino médio completo"),
    ("i", "2. Experiência mínima de 1 ano em produção gráfica"),
    ("i", "3. Disponibilidade para trabalhar em turnos"),
    ("i", "4. Residir em Fortaleza ou região metropolitana"),
    ("", ""),
    ("h", "Requisitos desejáveis"),
    ("i", "CNH categoria B"),
    ("i", "Curso técnico em artes gráficas"),
    ("i", "Experiência com impressora Heidelberg"),
    ("", ""),
    ("h", "Atividades"),
    ("i", "Auxiliar na preparação e no acerto da máquina offset"),
    ("i", "Conferir tiragem, registro e cor contra a prova aprovada"),
    ("i", "Abastecer papel e tinta e registrar o consumo na OS"),
    ("i", "Zelar pela limpeza e pela manutenção de primeiro nível"),
    ("", ""),
    ("h", "Observações"),
    ("i", "Processo com triagem cega: as inscrições são identificadas por código."),
    ("i", "Inscrições recebidas entre 01/07/2026 e 14/07/2026."),
]
for r, (tipo, txt) in enumerate(LINHAS_VAGA, start=2):
    cel = wv.cell(row=r, column=2, value=txt)
    if tipo == "t":
        cel.font = Font(bold=True, size=14)
    elif tipo == "s":
        cel.font = Font(size=10, italic=True)
    elif tipo == "h":
        cel.font = Font(bold=True, size=11)
        cel.fill = PatternFill("solid", fgColor="EEEEEE")

destino = ("/Users/rafaellima/developer/4-cursos-treinamentos/treinamentos-in-company/"
           "pouchain-claude-na-pratica/site/m1/a2-pedir-para-entregar/exercicio/"
           "inscricoes-vaga-auxiliar-offset.xlsx")
wb.save(destino)

# ── os fatos que as páginas vão afirmar ─────────────────────────────────────
print(f"gravado em {destino}\n")
print(f"{'inscrições':38} {N}")
print(f"{'escolaridade escrita de N jeitos':38} {len(MEDIO)} (mesmo nível)")
print(f"{'células de CNH em branco':38} 3")
print(f"{'células de disponibilidade em branco':38} 2")

for estrito, rotulo in [(False, "LEITURA A · qualquer turno fixo serve"),
                        (True, "LEITURA B · precisa rodar entre os turnos")]:
    atende, quase, nao = [], [], []
    for p in insc:
        f = obrig(p, estrito)
        (atende if not f else quase if len(f) == 1 else nao).append((p, f))
    print(f"\n── {rotulo} ──")
    print(f"   atende a todos os obrigatórios   {len(atende):>3}")
    print(f"   atende a todos menos um          {len(quase):>3}")
    print(f"   não atende                       {len(nao):>3}")

a_frouxo = sum(1 for p in insc if not obrig(p, False))
a_estrito = sum(1 for p in insc if not obrig(p, True))
print(f"\nARMADILHA 1 · a mesma frase da vaga dá {a_frouxo} ou {a_estrito} "
      f"aprovados.\n   Diferença de {a_frouxo - a_estrito} pessoas, e nenhuma "
      f"das duas leituras é\n   errada. Por isso o pedido tem que mandar "
      f"perguntar antes de aplicar.")

des = sum(1 for p in insc if not obrig(p, False) and p["cnh"] in ("B", "AB")
          and p["tecnico"] == "Sim" and p["heidelberg"] == "Sim")
print(f"\nARMADILHA 2 · exigindo também os 3 desejáveis sobram {des} "
      f"de {a_frouxo}.")

t = [p for p in insc if p["gr"] == 0 and p["anos_ult"] >= 9]
print(f"\nARMADILHA 3 · {len(t)} inscritos têm 9 anos ou mais na última "
      f"empresa\n   e ZERO em produção gráfica. Ex.: "
      f"{', '.join(p['cod'] for p in t[:4])}")
