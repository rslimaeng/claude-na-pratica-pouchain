# -*- coding: utf-8 -*-
"""
Planilha da DEMONSTRACAO da aula 1.1 · Compras da Grafica Aurora.
104 linhas: 26 insumos x 4 fornecedores.

O tamanho e o argumento da aula. Em 15 linhas o aluno pensa "isso eu fazia na
mao"; em 104, com fornecedor incompleto e unidade trocada, fazer na mao deixa
de ser demorado e passa a ser inviavel.

Quatro armadilhas, e nenhuma aparece passando o olho:
  1. PRAZO      Insumos Delta e o mais barato por unidade e chega depois de a
                maquina rodar.
  2. FRETE      Delta e Cearapel cotam FOB: o frete fica fora do valor unitario.
  3. INCOMPLETO Suprimentos e Cearapel deixaram itens sem cotar. Comparar total
                contra total e comparar coisas diferentes.
  4. UNIDADE    Cearapel cota a chapa CTP violeta em CAIXA COM 10. Ler o numero
                direto multiplica aquele item por dez.

Sujeira estrutural: cabecalho mesclado, coluna com espaco no nome, data em tres
formatos, valor ora numero ora texto, linha em branco no meio, frete em quatro
grafias, celulas vazias, unidade inconsistente.
"""
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

random.seed(2481)
HOJE, ENTRA_MAQUINA = datetime(2026, 7, 16), datetime(2026, 7, 22)
FOLGA = (ENTRA_MAQUINA - HOJE).days

# (descricao, unidade, quantidade, preco base = o da Papelaria Norte)
ITENS = [
    ("Papel couché brilho 150g 66x96",  "resma", 120, 289.00),
    ("Papel couché fosco 170g 66x96",   "resma",  80, 324.00),
    ("Papel offset 90g 66x96",          "resma",  40, 172.50),
    ("Papel offset 120g 66x96",         "resma",  35, 198.00),
    ("Cartão duplex 250g 70x100",       "resma",  60, 412.00),
    ("Cartão triplex 300g 70x100",      "resma",  25, 487.00),
    ("Cartão supremo 300g 66x96",       "resma",  45, 465.00),
    ("Papel kraft 80g bobina",          "kg",    200,   9.80),
    ("Tinta escala cyan",               "kg",     18, 148.00),
    ("Tinta escala magenta",            "kg",     18, 151.00),
    ("Tinta escala amarela",            "kg",     18, 146.00),
    ("Tinta escala preta",              "kg",     30, 132.00),
    ("Tinta Pantone 485C",              "kg",     12, 214.00),
    ("Tinta Pantone 072C",              "kg",      9, 228.00),
    ("Verniz UV brilho",                "litro",  25,  96.00),
    ("Verniz acrílico fosco",           "litro",  15,  88.50),
    ("Verniz de sobreimpressão",        "litro",  20,  79.00),
    ("Chapa CTP violeta 745x605",       "un",     60,  41.50),
    ("Chapa CTP térmica 1030x790",      "un",     45,  68.00),
    ("Blanqueta offset 1030x790",       "un",      8, 742.00),
    ("Filme BOPP laminação brilho",     "m2",   3000,   1.85),
    ("Filme BOPP laminação fosco",      "m2",   1800,   2.14),
    ("Cola hot melt",                   "kg",     40,  27.50),
    ("Solvente de limpeza",             "litro",  60,  18.90),
    ("Wire-o preto 3/4",                "un",    500,   3.40),
    ("Fita dupla face 12mm",            "rolo",   60,  22.80),
]

# fornecedor -> (multiplicador, prazo, grafia do frete, valor do frete,
#                condicao de pagamento, data da cotacao)
FORN = {
    "Papelaria Norte":        (1.000,  4, "CIF",    0.0, "28 dias",            "13/07/2026"),
    "Insumos Delta":          (0.938, 12, "fob",  890.0, "45 dias",            "13-07-26"),
    "Gráfica Suprimentos SA": (0.976,  6, "C.I.F.", 0.0, "à vista (3% desc.)", "2026-07-13"),
    "Distribuidora Cearapel": (0.957,  3, "F.O.B.", 1240.0, "21 dias",         "14/07/2026"),
}

# indices (base 0) que cada fornecedor deixou de cotar
SEM_COTAR = {"Gráfica Suprimentos SA": {2, 5, 15, 19, 24},
             "Distribuidora Cearapel": {10, 18, 22}}
# item 17 = Chapa CTP violeta. A Cearapel cotou em caixa fechada de 10
CAIXA = ("Distribuidora Cearapel", 17, "cx c/ 10", 10)

# ── monta a matriz de precos ────────────────────────────────────────────────
preco = {}                     # (fornecedor, idx) -> valor unitario cotado
unidade = {}                   # (fornecedor, idx) -> unidade cotada
for idx, (desc, un, qtd, base) in enumerate(ITENS):
    for f, (mult, *_r) in FORN.items():
        if idx in SEM_COTAR.get(f, ()):
            preco[(f, idx)] = None
            unidade[(f, idx)] = un
            continue
        ruido = random.uniform(0.985, 1.015)
        v = round(base * mult * ruido, 2)
        if (f, idx) == (CAIXA[0], CAIXA[1]):
            v = round(v * CAIXA[3], 2)          # preco da caixa, nao da unidade
            unidade[(f, idx)] = CAIXA[2]
        else:
            unidade[(f, idx)] = un
        preco[(f, idx)] = v

# ── escreve o arquivo ───────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Mapa de cotacao"

ws.merge_cells("A1:J1")
ws["A1"] = "GRÁFICA AURORA · SUPRIMENTOS"
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:J2")
ws["A2"] = ("Mapa de cotação · OS 2481 · catálogo institucional 20.000 ex. · "
            "exportado em 14/07/2026 17:38")
ws["A2"].font = Font(size=10, italic=True)
ws["A2"].alignment = Alignment(horizontal="center")

COLS = ["Item", "Descrição", "Unid", "Qtd", "Fornecedor", "Valor unit",
        " Prazo entrega ", "Frete", "Cond. pagamento", "Data cotação"]
for i, c in enumerate(COLS, start=1):
    cel = ws.cell(row=4, column=i, value=c)
    cel.font = Font(bold=True, size=10)
    cel.fill = PatternFill("solid", fgColor="DDDDDD")


def moeda(v):
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


linha, n_texto, n_vazias = 5, 0, 0
for idx, (desc, un, qtd, base) in enumerate(ITENS):
    for f, (mult, prazo, frete, vfrete, pag, data) in FORN.items():
        p = preco[(f, idx)]
        if p is None:
            n_vazias += 1
        texto = (linha % 3 == 0) and p is not None
        if texto:
            n_texto += 1
        ws.cell(row=linha, column=1, value=idx + 1)
        ws.cell(row=linha, column=2, value=desc)
        ws.cell(row=linha, column=3, value=unidade[(f, idx)])
        ws.cell(row=linha, column=4, value=qtd)
        ws.cell(row=linha, column=5, value=f)
        ws.cell(row=linha, column=6, value=moeda(p) if texto else p)
        ws.cell(row=linha, column=7, value=prazo)
        ws.cell(row=linha, column=8, value=frete)
        ws.cell(row=linha, column=9, value=pag)
        ws.cell(row=linha, column=10, value=data)
        linha += 1
    if idx == 12:                       # linha em branco no meio do arquivo
        linha += 1

linha += 1
for nota in ["Frete FOB por conta da Aurora: Insumos Delta R$ 890,00 · "
             "Distribuidora Cearapel R$ 1.240,00",
             "Produção da OS 2481 entra em máquina em 22/07/2026",
             "Cotações válidas por 5 dias corridos a partir da data de envio"]:
    ws.cell(row=linha, column=1, value=nota).font = Font(italic=True, size=9)
    linha += 1

for i, w in enumerate([6, 34, 10, 6, 24, 12, 15, 8, 20, 14], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

RAIZ = ("/Users/rafaellima/developer/4-cursos-treinamentos/treinamentos-in-company/"
        "pouchain-claude-na-pratica/site/m1/")
# o mesmo arquivo serve a demonstração da 1.1 e o caso de Compras do exercício
# da 1.2. Duas cópias saem do MESMO gerador, de propósito: assim não divergem.
DESTINOS = [RAIZ + "a1-ecossistema-e-fisica/demonstracao/cotacoes-fornecedores.xlsx",
            RAIZ + "a2-pedir-para-entregar/exercicio/cotacoes-fornecedores.xlsx"]
for destino in DESTINOS:
    wb.save(destino)

# ── os fatos que as páginas vão afirmar ─────────────────────────────────────
print("gravado em:")
for d in DESTINOS:
    print("  " + d)
print()
print(f"{'linhas de cotação':30} {len(ITENS) * len(FORN)}")
print(f"{'insumos × fornecedores':30} {len(ITENS)} × {len(FORN)}")
print(f"{'valores gravados como texto':30} {n_texto}")
print(f"{'células de preço vazias':30} {n_vazias}")
print(f"{'folga até a máquina rodar':30} {FOLGA} dias\n")

resumo = {}
for f, (mult, prazo, frete, vfrete, pag, data) in FORN.items():
    sub = 0.0
    for idx, (desc, un, qtd, base) in enumerate(ITENS):
        p = preco[(f, idx)]
        if p is None:
            continue
        sub += (qtd / CAIXA[3]) * p if (f, idx) == (CAIXA[0], CAIXA[1]) else p * qtd
    real = sub + vfrete
    if "vista" in pag:
        real = round(real * 0.97, 2)
    falt = sorted(SEM_COTAR.get(f, ()))
    resumo[f] = dict(real=real, prazo=prazo, faltam=len(falt),
                     atraso=max(0, prazo - FOLGA))
    marca = f"ATRASA {prazo - FOLGA}d" if prazo > FOLGA else "no prazo"
    print(f"{f:24} total real {real:>12,.2f} | {prazo:>2}d {marca:<10} | "
          f"faltam {len(falt)} itens")
    if falt:
        print(f"{'':24} não cotou: {', '.join(ITENS[i][0] for i in falt)}")

f_cx, i_cx = CAIXA[0], CAIXA[1]
naive = preco[(f_cx, i_cx)] * ITENS[i_cx][2]
certo = (ITENS[i_cx][2] / CAIXA[3]) * preco[(f_cx, i_cx)]
print(f"\narmadilha da caixa: {f_cx} cotou '{ITENS[i_cx][0]}' em {CAIXA[2]}")
print(f"   ler direto: {naive:,.2f}   correto: {certo:,.2f}   "
      f"erro de {naive - certo:,.2f}")

viaveis = [f for f, d in resumo.items() if d["faltam"] == 0 and d["atraso"] == 0]
barato  = min(resumo, key=lambda f: resumo[f]["real"])
print(f"\ncotou tudo E entrega a tempo: {viaveis}")
print(f"mais barato na conta ingênua: {barato} "
      f"(faltam {resumo[barato]['faltam']} itens)")
print(f"diferença entre o 'mais barato' e o único viável: "
      f"{resumo[viaveis[0]]['real'] - resumo[barato]['real']:,.2f}")
