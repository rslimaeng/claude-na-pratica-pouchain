#!/usr/bin/env python3
"""
Gates de qualidade do site · rode da raiz do repo:  python3 goals/gates.py

Cada onda fecha rodando isto. Sai com código 1 se algum gate falhar, então
serve como hook de pre-commit se um dia a gente quiser.

Regra que vale para escrever gate novo aqui:
  1. Gate com exceção permanente deixa de ser gate. Se precisa de exceção,
     conserta o código, não o gate.
  2. Gate que procura uma palavra NÃO PODE rodar contra o arquivo que enuncia
     a regra sobre aquela palavra, senão se auto-reprova. Ver EXCLUI_REGRA.
"""
import re, glob, os, sys
from html.parser import HTMLParser

EM = chr(0x2014)                       # o travessão, escrito por escape de propósito
AULAS = sorted(glob.glob("m1/a*/index.html"))
HTML  = sorted(glob.glob("*.html")) + sorted(glob.glob("m1/**/*.html", recursive=True))
MD    = sorted(glob.glob("*.md")) + sorted(glob.glob("**/*.md", recursive=True))
# arquivos que ENUNCIAM as regras. Procurar a palavra proibida aqui é falso positivo.
EXCLUI_REGRA = {"CLAUDE.md", "goals/goal-01-infra-e-padrao.md", "goals/gates.py",
                "goals/README.md", "README.md", "_shared/design-tokens.md"}

falhas = []
def diz(cond, gate, alvo="", extra=""):
    print(f"  {'OK   ' if cond else 'FALHA'} {alvo:44}{extra}")
    if not cond: falhas.append(f"{gate} {alvo}".strip())

def titulo(t): print("\n" + t)

# ─────────────────────────────────────────────────── G1 · anatomia da aula
titulo("G1 · ESTRUTURA · os blocos obrigatórios da página de aula")
MARC = ["o que você vai saber fazer", "a situação", "o conceito", "como funciona",
        "demonstração", "sua vez", "confira", "pegadinhas", "a cerca",
        "confira você mesmo",                       # o validador de fim de aula
        "checkpoint", "cobertura da ementa"]
for f in AULAS:
    s = open(f, encoding="utf-8").read().lower()
    falta = [m for m in MARC if m not in s]
    # o hook fecha toda aula. Na última do módulo ele fecha o módulo, e vale igual.
    if not re.search(r"a trava que est[ae] (aula|módulo) não resolve", s):
        falta.append("hook")
    diz(not falta, "G1", f, f"{len(MARC)+1-len(falta)}/{len(MARC)+1}  {falta or ''}")

# ─────────────────────────────────────────────────── G2 · gabarito escondido
titulo("G2 · GABARITO ATRÁS DE <details>, NUNCA ABERTO")
for f in AULAS:
    s = open(f, encoding="utf-8").read()
    det = len(re.findall(r'<details class="gabarito"', s))
    ab  = len(re.findall(r'<details class="gabarito" open', s))
    diz(det >= 1 and ab == 0, "G2", f, f"details={det} open={ab}")

# ─────────────────────────────────────────────────── G3 · higiene de dados
titulo("G3 · HIGIENE · o site é público")
# 3a. Página nenhuma usa o nome real do cliente: o universo é a Gráfica Aurora.
for f in HTML:
    s = open(f, encoding="utf-8").read()
    diz(not re.search(r"\bPouchain\b", s), "G3a", f,
        "nome do cliente em página de conteúdo" if re.search(r"\bPouchain\b", s) else "")
# 3b. Achado de consultoria não entra em lugar nenhum, nem em .md de governança.
ACHADO = [r"\bshadow ai\b", r"\bLGPD\b", r"\d+\s+casos? confirmados?",
          r"fragilidade estrutural", r"risco confirmado", r"\bCPF\b\s*[:\d]"]
for f in HTML + MD:
    if f in EXCLUI_REGRA: continue
    s = open(f, encoding="utf-8").read()
    hits = [p for p in ACHADO if re.search(p, s, re.I)]
    diz(not hits, "G3b", f, str(hits) if hits else "")

# ─────────────────────────────────────────────────── G4 · fato de produto datado
titulo("G4 · AFIRMAÇÃO SOBRE PRODUTO PRECISA DE DATA DE VERIFICAÇÃO")
AFIRMA = r"Pro, Max|Team ou Enterprise|proprietário da conta|individual de cada|não sincroniza"
for f in AULAS:
    s = open(f, encoding="utf-8").read()
    n = len(re.findall(AFIRMA, s))
    d = len(re.findall(r"verificad[oa].{0,40}\d{2}/\d{2}/\d{4}", s, re.I))
    diz(n == 0 or d >= 1, "G4", f, f"afirmações={n} datas={d}")

# ─────────────────────────────────────────────────── G6 · navegação
titulo("G6 · NAVEGAÇÃO · link ou arquivo referenciado que não existe")
# `src` entra junto com `href`: a primeira imagem do site entrou em 07/08 e o
# gate olhava só href, então um caminho de imagem errado passaria em silêncio.
mortos, n_ref = [], 0
for f in HTML:
    base = os.path.dirname(f)
    src = open(f, encoding="utf-8").read()
    for attr, ref in re.findall(r'(href|src)="([^"#][^"]*)"', src):
        if ref.startswith(("http", "mailto", "data:")): continue
        # link para uma âncora DE OUTRA página (`../#s06`) é legítimo, e o gate
        # reprovava porque tentava achar um arquivo chamado "#s06". A regex já
        # ignora href que COMEÇA com `#`; faltava cortar o fragmento do resto.
        ref = ref.split("#", 1)[0]
        if not ref: continue
        n_ref += 1
        alvo = os.path.normpath(os.path.join(base, ref))
        if os.path.isdir(alvo): alvo = os.path.join(alvo, "index.html")
        if not os.path.exists(alvo): mortos.append(f"{f} -> {attr}={ref}")
diz(not mortos, "G6", "todos os links e arquivos",
    str(mortos) if mortos else f"{n_ref} referências, 0 morta")

# toda imagem precisa de alt, e de width/height para não pular o layout ao carregar
sem = []
for f in HTML:
    for tag in re.findall(r"<img\b[^>]*>", open(f, encoding="utf-8").read()):
        # o atributo tem que vir depois de ESPAÇO. `data-alt="` contém `alt="`
        # como substring, e `\b` não resolve porque o hífen já é fronteira de
        # palavra. Os dois furos apareceram testando o gate contra defeito
        # injetado, não lendo o código.
        falta = [a for a in ("alt", "width", "height")
                 if not re.search(rf'\s{a}="', tag)]
        if falta: sem.append(f"{f} {falta}")
diz(not sem, "G6b", "toda <img> com alt, width e height",
    str(sem) if sem else "")

# ─────────────────────────────────────────────────── G7 · render
titulo("G7 · RENDER · balanço de tags")
class Bal(HTMLParser):
    VOID = {"meta","link","br","img","hr","input","source","col"}
    def __init__(s): super().__init__(); s.p=[]; s.err=[]
    def handle_starttag(s,t,a):
        if t not in s.VOID: s.p.append(t)
    def handle_endtag(s,t):
        if not s.p: s.err.append(f"</{t}> sobrando"); return
        if s.p[-1]==t: s.p.pop()
        elif t in s.p:
            while s.p and s.p[-1]!=t: s.err.append(f"<{s.p.pop()}> não fechada")
            s.p.pop()
for f in HTML:
    p = Bal(); p.feed(open(f, encoding="utf-8").read())
    diz(not p.err and not p.p, "G7", f, str((p.err+p.p)[:3]) if (p.err or p.p) else "")

# ─────────────────────────────────────── G7-ter · classe usada sem CSS
titulo("G7-ter · CLASSE USADA SEM NENHUM CSS QUE A PEGUE")
for f in HTML:
    src = open(f, encoding="utf-8").read()
    style = "\n".join(re.findall(r"<style>(.*?)</style>", src, re.S))
    body  = re.sub(r"<style>.*?</style>", "", src, flags=re.S)
    usadas = set()
    for m in re.findall(r'class="([^"]+)"', body): usadas.update(m.split())
    falta = sorted(usadas - set(re.findall(r"\.([A-Za-z][\w-]*)", style)))
    diz(not falta, "G7ter", f, str(falta) if falta else "")

# ─────────────────────────────────────────────────── G9 · travessão
titulo("G9 · TRAVESSÃO")
tot = {f: open(f, encoding="utf-8").read().count(EM) for f in HTML + MD}
tot = {f: n for f, n in tot.items() if n}
diz(not tot, "G9", "todo o repo", str(tot) if tot else f"0 em {len(HTML)+len(MD)} arquivos")

# ─────────────────────────────────────────────────── G10 · formato de insumo
titulo("G10 · FORMATO DOS INSUMOS")
ruins = glob.glob("m1/**/*.csv", recursive=True) + glob.glob("m1/**/*.txt", recursive=True)
ins = sorted(glob.glob("m1/**/exercicio/*", recursive=True) +
             glob.glob("m1/**/gabarito/*", recursive=True))
for i in ins: print(f"        {i}")
diz(not ruins, "G10", "insumos", f"{len(ins)} arquivo(s), {len(ruins)} em formato proibido")

# ─────────────────────────────────────── G11 · inline forçada a bloco
titulo("G11 · TAG INLINE FORÇADA A display:block")
print("        (quebra a frase no meio quando a tag aparece dentro de um <p>)")
INLINE = {"strong","em","b","i","span","a","code","small","abbr"}
for f in HTML:
    style = "\n".join(re.findall(r"<style>(.*?)</style>",
                                 open(f, encoding="utf-8").read(), re.S))
    bad = []
    for sel, b in re.findall(r"([^{}@]+)\{([^}]*)\}", style):
        if "display:block" not in b.replace(" ", "").replace("\n", ""): continue
        sel = sel.strip()
        if sel.split(",")[0].split()[-1].strip() in INLINE and ">" not in sel:
            bad.append(sel)
    diz(not bad, "G11", f, str(bad) if bad else "")

# ─────────────────────────────────── G11b · flex/grid engolindo frase
titulo("G11b · CONTAINER flex/grid COM TEXTO SOLTO + TAG INLINE")
print("        (cada pedaço vira item de flex e a frase quebra no meio)")
INLINE_T = {"strong","em","b","i","span","a","code","small","abbr"}

class Flexy(HTMLParser):
    """Acha elemento flex/grid que tem, como filho DIRETO, texto solto e tag inline.
    Regex não serve aqui: com <div> dentro de <div> ela fecha na tag errada."""
    VOID = {"meta","link","br","img","hr","input","source","col"}
    def __init__(s, alvo):
        super().__init__(); s.alvo = alvo; s.pilha = []; s.achados = set()
    def handle_starttag(s, t, attrs):
        cls = dict(attrs).get("class", "").split()
        marcado = next((c for c in cls if c in s.alvo), None)
        if s.pilha and t in INLINE_T: s.pilha[-1]["inline"] = True
        if t not in s.VOID:
            s.pilha.append({"tag": t, "cls": marcado, "texto": 0, "inline": False})
    def handle_endtag(s, t):
        while s.pilha:
            n = s.pilha.pop()
            if n["cls"] and n["texto"] > 12 and n["inline"]: s.achados.add(n["cls"])
            if n["tag"] == t: break
    def handle_data(s, d):
        if s.pilha: s.pilha[-1]["texto"] += len(d.strip())

for f in HTML:
    src = open(f, encoding="utf-8").read()
    style = "\n".join(re.findall(r"<style>(.*?)</style>", src, re.S))
    body  = re.sub(r"<style>.*?</style>", "", src, flags=re.S)
    alvo = set()
    for sel, b in re.findall(r"([^{}@]+)\{([^}]*)\}", style):
        bb = b.replace(" ", "").replace("\n", "")
        if "display:flex" in bb or "display:grid" in bb:
            u = sel.strip().split(",")[0].split()[-1].strip()
            if u.startswith("."): alvo.add(u[1:].split(":")[0])
    fp = Flexy(alvo); fp.feed(body)
    diz(not fp.achados, "G11b", f, str(sorted(fp.achados)) if fp.achados else "")

# ─────────────── G11c · o atalho `margin:` que mata a centralização do bloco largo
titulo("G11c · BLOCO LARGO COM `margin:` DEPOIS DA REGRA DE SANGRAMENTO")
print("        (o atalho zera o margin-left:50% e o bloco vai 570px para a esquerda)")
# A regra de sangramento centra o bloco largo com margin-left:50% +
# transform:translateX(-50%). Qualquer regra POSTERIOR que use o atalho
# `margin:` zera o margin-left e sobra só o transform, que empurra o bloco
# meia largura para a esquerda. O .compare escapava por sorte: a regra dele
# vem ANTES do sangramento. Defeito invisível em grep de conteúdo, e só
# aparece em tela larga, porque em 1280 a coluna quase encosta na borda.
for f in HTML:
    css = "\n".join(re.findall(r"<style>(.*?)</style>",
                               open(f, encoding="utf-8").read(), re.S))
    css = re.sub(r"/\*.*?\*/", "", css, flags=re.S)   # comentário entra no seletor
    regras = [(m.start(), m.group(1).strip(), re.sub(r"\s+", " ", m.group(2)))
              for m in re.finditer(r"([^{}@]+)\{([^}]*)\}", css)]
    sangra = [(p, s) for p, s, b in regras if "margin-left:50%" in b.replace(" ", "")]
    if not sangra:
        diz(True, "G11c", f, "sem regra de sangramento"); continue
    pos_s, sel_s = sangra[0]
    largos = {c.strip().lstrip(".") for c in sel_s.split(",") if c.strip().startswith(".")}
    ruins = sorted({s.strip() for p, s, b in regras
                    if p > pos_s and s.strip().lstrip(".") in largos
                    and re.search(r"(^|;)\s*margin\s*:", b)})
    diz(not ruins, "G11c", f,
        f"atalho margin: em {ruins}" if ruins else f"{len(largos)} blocos largos, 0 quebrado")

# ─────────────────────── G12 · o nome do nível é o aprovado, e vem com o recurso
titulo("G12 · NOME DE NÍVEL APROVADO + RECURSO OFICIAL DO CLAUDE AO LADO")
print("        (vocabulário interno não vai para a tela · ver CLAUDE.md §5 e §7-bis)")
NOMES_OK = ["Por que ele piora na conversa longa",
            "Pedir uma vez e receber pronto",
            "Ele já começa sabendo as suas regras",
            "Cada tarefa puxa o seu próprio procedimento",
            "Ele abre os arquivos onde você trabalha",
            "Você prova que está certo antes de mandar",
            "Roda sem você apertar o play"]
land = open("index.html", encoding="utf-8").read()
achados = re.findall(r'class="nivel-nome">([^<]+)<', land)
diz(achados == NOMES_OK, "G12", "index.html · nomes da trilha",
    "" if achados == NOMES_OK else f"divergem de CLAUDE.md §5: {set(achados) ^ set(NOMES_OK)}")
# toda linha da trilha mostra pelo menos um recurso oficial
linhas = land.split('class="nivel-row')[1:]
sem = [i for i, b in enumerate(linhas) if 'class="rec' not in b.split("</div>\n      </div>")[0]]
diz(not sem, "G12", "index.html · recurso em toda linha",
    "" if not sem else f"linhas sem .rec: {sem}")
# hero de aula e card do hub carregam o recurso
for f in AULAS + ["m1/index.html"]:
    src = open(f, encoding="utf-8").read()
    n = len(re.findall(r'class="rec[" ]', src))
    esperado = 4 if f.endswith("m1/index.html") else 1
    diz(n >= esperado, "G12", f, f"chips de recurso={n} (mínimo {esperado})")

# ─────────────── G13 · número sobre a planilha tem que sair da planilha
titulo("G13 · OS NÚMEROS DA GRÁFICA AURORA BATEM COM O .XLSX")
print("        (nenhum número sobre insumo é digitado à mão: tudo recalculado aqui)")
try:
    from openpyxl import load_workbook
    from datetime import datetime, timedelta
    HOJE = datetime(2026, 7, 16)
    MES = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
           "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}

    def data(v):
        if isinstance(v, datetime): return v
        s = str(v).strip()
        for f in ("%d/%m/%Y", "%Y-%m-%d"):
            try: return datetime.strptime(s, f)
            except ValueError: pass
        p = s.split("-")
        if len(p) == 3 and p[1] in MES:
            return datetime(2000 + int(p[2]), MES[p[1]], int(p[0]))
        return None

    def texto(p):
        return re.sub(r"<[^>]+>", " ", open(p, encoding="utf-8").read()).lower()

    N = {1:"um",2:"dois",3:"três",4:"quatro",5:"cinco",6:"seis",8:"oito",
         9:"nove",10:"dez",12:"doze",14:"quatorze"}

    # ── PLANILHA DO EXERCÍCIO · PCP, 120 OS ────────────────────────────────
    ws = load_workbook("m1/a1-ecossistema-e-fisica/exercicio/"
                       "pedidos-em-producao.xlsx", data_only=True).active
    linhas = [r for r in ws.iter_rows(min_row=5, values_only=True)
              if r[0] and str(r[0]).isdigit()]
    vivas = [r for r in linhas if str(r[6]).strip().lower() != "entregue"]
    atras = [r for r in vivas if (d := data(r[5])) and d < HOJE]
    marcadas = [r for r in linhas if str(r[6]).strip() == "Atrasado"]
    todas_os = [r[0] for r in linhas]
    dups = sorted({o for o in todas_os if todas_os.count(o) > 1})
    print(f"        {'PCP · linhas':28} {len(linhas)}")
    print(f"        {'PCP · em aberto':28} {len(vivas)}")
    print(f"        {'PCP · atrasadas pela data':28} {len(atras)}")
    print(f"        {'PCP · marcadas no sistema':28} {len(marcadas)}   (a mentira do campo)")
    print(f"        {'PCP · OS repetidas':28} {dups}")

    ex   = texto("m1/a1-ecossistema-e-fisica/exemplo/index.html")
    a1   = texto("m1/a1-ecossistema-e-fisica/index.html")
    a3   = texto("m1/a3-regra-que-fica/index.html")
    a3d  = texto("m1/a3-regra-que-fica/demonstracao/index.html")
    a2d  = texto("m1/a2-pedir-para-entregar/demonstracao/index.html")

    for ok, nome in [
        (f">{len(vivas)}<" in open("m1/a1-ecossistema-e-fisica/exemplo/index.html",
                                   encoding="utf-8").read(), "PCP · exemplo · em produção"),
        (f'alerta">{len(atras)}<' in open("m1/a1-ecossistema-e-fisica/exemplo/index.html",
                                          encoding="utf-8").read(), "PCP · exemplo · atrasadas"),
        (f"{len(linhas)} linhas" in ex,            "PCP · exemplo · tamanho do arquivo"),
        (f"{len(dups)} os repetidas" in ex,        "PCP · exemplo · OS duplicadas"),
        (f"{len(linhas)} ordens de serviço" in a1, "PCP · aula 1.1 · tamanho no download"),
        (f"{len(linhas)} ordens de serviço" in a2d,"PCP · aula 1.2 · tamanho no download"),
        (f"{len(atras)} atrasadas" in a3,          "PCP · aula 1.3 · atrasadas"),
        (f"marca <strong>{len(marcadas)}</strong>" in
         open("m1/a3-regra-que-fica/index.html", encoding="utf-8").read(),
                                                   "PCP · aula 1.3 · marcadas no sistema"),
        (f"{len(atras)} os atrasadas" in a3d,      "PCP · demo 1.3 · atrasadas"),
        (f"marca <b>{len(marcadas)}</b>" in
         open("m1/a3-regra-que-fica/demonstracao/index.html", encoding="utf-8").read(),
                                                   "PCP · demo 1.3 · marcadas no sistema"),
    ]: diz(bool(ok), "G13", nome)

    # ── PLANILHA DA DEMONSTRAÇÃO · Compras, 104 linhas ─────────────────────
    wc = load_workbook("m1/a1-ecossistema-e-fisica/demonstracao/"
                       "cotacoes-fornecedores.xlsx", data_only=True).active
    FOLGA = (datetime(2026, 7, 22) - HOJE).days
    cot = [r for r in wc.iter_rows(min_row=5, values_only=True) if r[0] and r[4]]
    forn = {}
    for r in cot:
        f = forn.setdefault(r[4], {"prazo": r[6], "sem": 0, "caixa": None})
        if r[5] in (None, ""):        f["sem"] += 1
        if "cx" in str(r[2]).lower(): f["caixa"] = str(r[2])
    itens   = {r[0] for r in cot}
    atrasa  = {f: d["prazo"] - FOLGA for f, d in forn.items() if d["prazo"] > FOLGA}
    pior    = max((d["sem"] for d in forn.values()), default=0)
    caixa   = next((d["caixa"] for d in forn.values() if d["caixa"]), None)
    por_cx  = int(re.search(r"\d+", caixa).group()) if caixa else 0
    viaveis = [f for f, d in forn.items() if not d["sem"] and d["prazo"] <= FOLGA]
    print(f"        {'cotações · linhas':28} {len(cot)}")
    print(f"        {'cotações · itens × forn.':28} {len(itens)} × {len(forn)}")
    print(f"        {'cotações · maior atraso':28} {max(atrasa.values(), default=0)} dias")
    print(f"        {'cotações · itens sem cotar':28} {pior} (pior fornecedor)")
    print(f"        {'cotações · cotou tudo E no prazo':28} {viaveis}")

    d1  = texto("m1/a1-ecossistema-e-fisica/demonstracao/index.html")
    for ok, nome in [
        (f"{len(cot)} linhas de cotação" in d1 and f"{len(cot)} linhas" in a1,
         "cotações · tamanho da planilha"),
        (f"{len(forn)} fornecedores" in d1 and f"{len(forn)} fornecedores" in a1,
         "cotações · quantos fornecedores"),
        (f"{len(itens)} insumos" in d1, "cotações · quantos insumos"),
        (len(atrasa) == 1 and f"{N[list(atrasa.values())[0]]} dias" in d1,
         "cotações · de quantos dias é o atraso"),
        (pior and f"cotar {N[pior]} itens" in d1, "cotações · itens sem cotar"),
        (por_cx and f"caixa com {N[por_cx]}" in d1, "cotações · a armadilha da caixa"),
        (len(viaveis) == 1 and "um único fornecedor" in d1,
         "cotações · sobra um fornecedor só"),
    ]: diz(bool(ok), "G13", nome)

    # ── PLANILHAS DO EXERCÍCIO DA 1.2 ──────────────────────────────────────
    # elas nasceram porque o exercício mandava rodar no Claude e contar
    # mensagens, e não existia arquivo nenhum para anexar.
    EX2 = "m1/a2-pedir-para-entregar/exercicio/"
    fec = load_workbook(EX2 + "fechamento-dois-meses.xlsx", data_only=True)
    n_fec = sum(len([r for r in fec[a].iter_rows(min_row=5, values_only=True)
                     if r[0]]) for a in fec.sheetnames)
    ins = load_workbook(EX2 + "inscricoes-vaga-auxiliar-offset.xlsx",
                        data_only=True)["Inscrições"]
    n_ins = len([r for r in ins.iter_rows(min_row=5, values_only=True) if r[0]])
    print(f"        {'fechamento · linhas de conta':28} {n_fec}")
    print(f"        {'fechamento · abas':28} {fec.sheetnames}")
    print(f"        {'inscrições · linhas':28} {n_ins}")

    a2 = texto("m1/a2-pedir-para-entregar/index.html")
    for ok, nome in [
        (f"{n_fec} linhas" in a2, "1.2 · tamanho do fechamento"),
        (f"{n_ins} inscrições" in a2, "1.2 · quantas inscrições"),
        (f"{len(cot)} linhas" in a2, "1.2 · tamanho do mapa de cotação"),
        (f"{len(itens)} insumos" in a2, "1.2 · quantos insumos"),
        (all(os.path.exists(EX2 + x) for x in
             ["cotacoes-fornecedores.xlsx", "fechamento-dois-meses.xlsx",
              "inscricoes-vaga-auxiliar-offset.xlsx"]),
         "1.2 · os três casos têm arquivo"),
        # sem nome de pessoa: a triagem é por código, e material de curso
        # não carrega nome de gente em lugar nenhum
        (all(str(r[0]).startswith("INSC-")
             for r in ins.iter_rows(min_row=5, values_only=True) if r[0]),
         "1.2 · inscrições anonimizadas por código"),
    ]: diz(bool(ok), "G13", nome)

    # ── AS MÁQUINAS DA AURORA · o gabarito dizia 4, a planilha tem 6 ───────
    # Não era erro cosmético: o exercício da 1.3 manda colar essas regras num
    # Project e rodar na MESMA planilha. Gabarito com 4 máquinas ensina que
    # Digital 1 e Offset 3 não existem, e a aula que promete melhorar a
    # resposta entrega uma pior. Insumo é a fonte da verdade, texto obedece.
    import zipfile
    pcp = load_workbook("m1/a1-ecossistema-e-fisica/exercicio/"
                        "pedidos-em-producao.xlsx", data_only=True).active
    cab = list(next(pcp.iter_rows(min_row=4, max_row=4, values_only=True)))
    i_maq = [i for i, c in enumerate(cab) if c and "quina" in str(c)][0]
    maqs = sorted({r[i_maq] for r in linhas if r[i_maq]})
    print(f"        {'PCP · máquinas na planilha':28} {len(maqs)}  {maqs}")

    def docx_txt(p):
        x = zipfile.ZipFile(p).read("word/document.xml").decode("utf-8")
        return re.sub(r"<[^>]+>", "", x.replace("</w:p>", "\n"))

    gab = docx_txt("m1/a3-regra-que-fica/gabarito/minhas-regras-GABARITO.docx")
    a1d = texto("m1/a1-ecossistema-e-fisica/demonstracao/index.html")
    for alvo, nome in [(a1, "aula 1.1"), (a3, "aula 1.3"), (a3d, "demo 1.3"),
                       (gab.lower(), "gabarito 1.3")]:
        falta = [m for m in maqs if m.lower() not in alvo]
        cita = "máquina" in alvo or "offset" in alvo
        diz(not (cita and falta), "G13", f"máquinas · {nome}",
            f"não cita {falta}" if falta else f"as {len(maqs)}")
    # a varredura roda em TODA página do aluno, não nas duas que eu lembrei:
    # o "quatro máquinas" da 1.4 e o "18 pedidos" da 1.1 escaparam justamente
    # de uma lista curta escrita à mão.
    TODAS = [(f, texto(f)) for f in HTML] + [("gabarito 1.3", gab.lower())]
    erradas = [n for n, t in TODAS
               if re.search(r"\b(quatro|4|cinco|5|sete|7|oito|8) máquinas", t)]
    diz(not erradas, "G13", "máquinas · nenhum texto afirma outro total",
        str(erradas) if erradas else f"{len(TODAS)} arquivos varridos")

    # nem uma contagem por família diferente da real. "duas máquinas offset"
    # passava no teste do total e mentia do mesmo jeito.
    NUM = {"uma": 1, "duas": 2, "dois": 2, "três": 3, "quatro": 4,
           "cinco": 5, "seis": 6}
    FAM = {"offset": "Offset", "flexográficas": "Flexo", "flexográfica": "Flexo",
           "flexo": "Flexo", "digitais": "Digital", "digital": "Digital"}
    real = {f: len([m for m in maqs if m.startswith(f)])
            for f in ("Offset", "Flexo", "Digital")}
    print(f"        {'PCP · por família':28} {real}")
    padrao = re.compile(r"\b(" + "|".join(NUM) + r") (?:máquinas? )?(" +
                        "|".join(FAM) + r")\b")
    ruins = {n: [f"{q} {f}" for q, f in padrao.findall(t)
                 if NUM[q] != real[FAM[f]]] for n, t in TODAS}
    ruins = {n: v for n, v in ruins.items() if v}
    diz(not ruins, "G13", "máquinas por família",
        f"{ruins}, e o real é {real}" if ruins else f"real {real}")

    # o total de OS também não pode ser inventado em lugar nenhum
    n_os = len(linhas)
    inventados = [n for n, t in TODAS
                  if re.search(r"\b(\d{1,3}) pedidos\b", t) and
                  any(int(x) != n_os for x in re.findall(r"\b(\d{1,3}) pedidos\b", t))]
    diz(not inventados, "G13", f"total de OS · nenhum texto diz outro que {n_os}",
        str(inventados) if inventados else "")

    # ── piso de tamanho · CLAUDE.md §8-ter ─────────────────────────────────
    for nome, n in [("pedidos-em-producao", len(linhas)),
                    ("cotacoes-fornecedores", len(cot)),
                    ("fechamento-dois-meses", n_fec),
                    ("inscricoes-vaga-auxiliar-offset", n_ins)]:
        diz(n >= 100, "G13", f"piso de 100 linhas · {nome}", f"{n} linhas")
except ImportError:
    print("        openpyxl ausente, gate pulado")

# ─── G13b · o .md de regras é insumo também, e a página afirma o tamanho dele
titulo("G13b · O ARQUIVO DE REGRAS DO EXEMPLO BATE COM O QUE A PÁGINA DIZ")
print("        (a 1.3 promete 'entre 20 e 40 linhas'. O exemplo tem que caber nela)")
REG = "m1/a3-regra-que-fica/exemplo/regras-do-coordenador.md"
if os.path.exists(REG):
    linhas_md = [l.rstrip() for l in open(REG, encoding="utf-8")]
    # o que de fato seria colado nas Instruções: sem título, sem o aviso em
    # citação e sem linha vazia. É esse número que a aula promete.
    n_reg = len([l for l in linhas_md
                 if l.strip() and not l.startswith(">") and not l.startswith("#")])
    n_sec = len([l for l in linhas_md if l.startswith("## ")])
    diz(20 <= n_reg <= 40, "G13b", "o exemplo cabe na faixa que a aula pede",
        f"{n_reg} linhas de regra")
    diz(n_sec == 5, "G13b", "as cinco seções do gabarito estão lá",
        f"{n_sec} seções")
    # e nenhuma página pode afirmar um número de linhas diferente do real
    erradas = []
    for f in HTML:
        for n in re.findall(r"(\d{1,3}) linhas de regra", open(f, encoding="utf-8").read()):
            if int(n) != n_reg: erradas.append(f"{f} diz {n}, e são {n_reg}")
    diz(not erradas, "G13b", f"nenhuma página afirma outro número que {n_reg}",
        str(erradas) if erradas else "")
else:
    diz(False, "G13b", REG, "arquivo não existe")

# ─── G13c · o Project de verdade da página de exemplo é descrito por números
#      que vêm de dois arquivos reais. Se um deles mudar de tamanho, a página
#      mente. Mesma lição do G13: número que a página afirma, o gate recalcula.
titulo("G13c · OS NÚMEROS DO PROJECT DE VERDADE BATEM COM OS ARQUIVOS")
print("        (a página de exemplo cita 128 linhas, 2.030 linhas e 15 especialistas)")
FONTE = "../insumos/exemplos/m1-aula-3"
EXEMPLO = "m1/a3-regra-que-fica/exemplo/index.html"
if os.path.isdir(FONTE):
    pag = open(EXEMPLO, encoding="utf-8").read()
    inst = os.path.join(FONTE, "00-System_Instruction.md")
    base = os.path.join(FONTE, "02-brand-squad.md")
    diz(os.path.exists(inst) and os.path.exists(base), "G13c",
        "os dois arquivos de origem estão lá")
    if os.path.exists(inst) and os.path.exists(base):
        n_inst = len(open(inst, encoding="utf-8").read().splitlines())
        linhas_base = open(base, encoding="utf-8").read().splitlines()
        n_base = len(linhas_base)
        # o roster: cada especialista é um "### " antes da matriz de roteamento
        fim = next((i for i, l in enumerate(linhas_base)
                    if l.startswith("## ") and "ROUTING" in l.upper()), len(linhas_base))
        n_esp = len([l for i, l in enumerate(linhas_base)
                     if l.startswith("### ") and i < fim])
        # o número vai para a tela com ponto de milhar, do jeito pt-BR
        def br(n): return f"{n:,}".replace(",", ".")
        diz(f"{n_esp} especialistas" in pag, "G13c", "especialistas no roster",
            f"a página precisa dizer '{n_esp} especialistas'")
        outros = [n for n in re.findall(r"(\d{1,3}) especialistas", pag) if int(n) != n_esp]
        diz(not outros, "G13c", f"nenhum outro número de especialistas que {n_esp}",
            str(outros) if outros else "")
        # Contagem de linhas: presença sozinha não protege, porque o número
        # aparece em mais de um lugar e trocar UM deles passaria batido.
        # Então cada "N linhas" dentro dos rótulos do painel é conferido.
        ROTULOS = r'class="(?:fcard-meta|forma-nome|painel-meta)">([^<]*)'
        citados = [int(n.replace(".", ""))
                   for bloco in re.findall(ROTULOS, pag)
                   for n in re.findall(r"([\d.]+) linhas", bloco)]
        validos = {n_inst, n_base}
        errados = [n for n in citados if n not in validos]
        diz(not errados, "G13c", "todo número de linhas do painel bate com o arquivo",
            f"{errados} não são {sorted(validos)}" if errados
            else f"{len(citados)} citações conferidas")
        diz(set(citados) == validos, "G13c", "os dois arquivos aparecem com o tamanho certo",
            f"achei {sorted(set(citados))}, esperava {sorted(validos)}")
else:
    print(f"  PULADO  {FONTE} não existe nesta máquina, os 5 números não foram conferidos")

# ─── G16 · o nome do campo é o que aparece na tela do aluno, não o do help center
titulo("G16 · O CAMPO SE CHAMA COMO ESTÁ ESCRITO NA TELA")
print("        (é 'Contexto' no claude.ai em pt-BR · verificado em 07/08/2026)")
# só o que o aluno lê: páginas e os .md publicados dentro de m1/.
# Os goals descrevem a regra e citam o nome velho de propósito.
DO_ALUNO = HTML + sorted(glob.glob("m1/**/*.md", recursive=True))
VELHOS = ["Conhecimento do projeto", "conhecimento do projeto", "Project knowledge"]
achados = []
for f in DO_ALUNO:
    s = open(f, encoding="utf-8").read()
    for v in VELHOS:
        if v in s: achados.append(f"{f} diz '{v}'")
diz(not achados, "G16", "nenhuma página do aluno usa o nome antigo do campo",
    str(achados) if achados else f"{len(DO_ALUNO)} arquivos varridos")
# e a aula que ensina a anatomia precisa nomear os três blocos da tela
ANAT = "m1/a3-regra-que-fica/index.html"
s = open(ANAT, encoding="utf-8").read() if os.path.exists(ANAT) else ""
faltam = [b for b in ["Instruções", "Memória", "Contexto"]
          if f'class="tela-caixa-rot">{b}<' not in s]
diz(not faltam, "G16", "a tela da 1.3 mostra os três blocos com o nome da tela",
    f"faltam {faltam}" if faltam else "")

# ─── G17 · faixa de tamanho é andaime do primeiro exercício, nunca regra do campo.
#      Nasceu porque a tabela da 1.3 dizia "Curto. 20 a 40 linhas" enquanto a
#      página de exemplo, a um clique dali, mostrava um campo de Instruções com
#      128 linhas funcionando. Número como critério contradiz o exemplo real.
titulo("G17 · FAIXA DE LINHAS SEMPRE QUALIFICADA COMO PONTO DE PARTIDA")
print("        (o critério é 'toda linha vale sempre'. O tamanho é consequência)")
QUALIF = ["primeira", "primeiras", "ponto de partida", "não é teto",
          "a aula pede", "não é o tamanho máximo"]
# A qualificação tem que estar NO MESMO BLOCO, não "por perto". Uma janela de
# caracteres atravessa <p> e <li> vizinhos e deixa passar a faixa crua porque
# o parágrafo de cima qualificava outra coisa. Aconteceu ao testar este gate.
BLOCO = re.compile(r"</?(?:p|div|li|td|tr|h[1-6]|ul|ol|table)\b[^>]*>")
crus = []
for f in HTML:
    s = open(f, encoding="utf-8").read()
    cortes = [0] + [m.end() for m in BLOCO.finditer(s)] + [len(s)]
    for m in re.finditer(r"\b\d{1,3}\s+[ae]\s+\d{1,3}\s+linhas\b", s):
        ini = max(c for c in cortes if c <= m.start())
        fim = min(c for c in cortes if c >= m.end())
        if not any(q in s[ini:fim].lower() for q in QUALIF):
            crus.append(f"{f}: '{m.group(0)}' cru, sem qualificação no mesmo bloco")
diz(not crus, "G17", "toda faixa de linhas vem qualificada",
    str(crus) if crus else f"{len(HTML)} páginas varridas")
# e a tabela de gavetas não volta a ter linha de tamanho como critério
AULA13 = "m1/a3-regra-que-fica/index.html"
s13 = open(AULA13, encoding="utf-8").read() if os.path.exists(AULA13) else ""
diz("<strong>Tamanho certo</strong>" not in s13, "G17",
    "a tabela de gavetas não usa tamanho como critério",
    "voltou a linha 'Tamanho certo'" if "<strong>Tamanho certo</strong>" in s13 else "")
# o critério também não pode voltar em outra unidade. O item 1 do validador da
# 1.3 media "cabe em uma página", que é a mesma régua com outro nome, e o
# exemplo da própria aula tem 128 linhas de instrução funcionando.
POR_TAMANHO = ["uma página", "duas páginas", "no máximo", "não passa de",
               "curto o suficiente", "menos de"]
medindo = []
for f in AULAS:
    s = open(f, encoding="utf-8").read()
    for item in re.findall(r'<div class="checagem-o">(.*?)</div>', s, re.S):
        alvo = re.sub(r"<[^>]+>", "", item).lower()
        for p in POR_TAMANHO:
            if p in alvo: medindo.append(f"{f}: '{alvo.strip()}' mede por {p}")
diz(not medindo, "G17", "nenhum item do validador mede o texto por tamanho",
    str(medindo) if medindo else f"{len(AULAS)} validadores varridos")

# ─── G19 · bloco que sangra para a largura larga tem que USAR a largura, e
#      componente de prosa não sangra. O card `.saidas` estourava para 1140px
#      com o texto travado em 58ch: 563px de vazio à direita, nas quatro aulas.
#      E o sangramento não pode acontecer dentro de um card, que é o que jogava
#      a tabela do gabarito para fora da própria caixa.
titulo("G19 · SANGRAMENTO SÓ ONDE ELE É USADO")
print("        (card de navegação é prosa e vive na coluna de leitura · §8-quater)")
# acha a regra pelo que ela FAZ, não por onde ela está. Procurar a partir do
# comentário "/* BREAKOUT" parava no primeiro `{` depois dele, e na 1.3 esse
# primeiro `{` era o do `.uau{--col-wide:960px}` que eu tinha posto logo abaixo.
# Fatiar por `}` em vez de casar com regex: `([^{}]+)\{...` fazia backtracking
# quadrático e travava em arquivo de 90 KB.
def regras_que_sangram(css):
    saida = []
    for pedaco in css.split("}"):
        if "margin-left:50%" not in pedaco:
            continue
        seletor = pedaco.rsplit("{", 1)[0] if "{" in pedaco else ""
        saida.append(seletor)
    return saida

for f in AULAS + [x for x in HTML if x.endswith("m1/index.html")]:
    s = open(f, encoding="utf-8").read()
    listas = regras_que_sangram(s)
    if not listas:
        continue
    sangrando = [l for l in listas if re.search(r"\.saidas\b", l)]
    diz(not sangrando, "G19", f,
        "card de navegação está sangrando" if sangrando
        else f"{len(listas)} regra(s) de sangramento, sem .saidas")
# tabela dentro do gabarito não pode sair da caixa
for f in AULAS:
    s = open(f, encoding="utf-8").read()
    corpos = re.findall(r'<details class="gabarito".*?</details>', s, re.S)
    tem_tabela = any("table-wrap" in c for c in corpos)
    if not tem_tabela:
        continue
    diz(".gabarito-body .table-wrap{" in s, "G19", f,
        "tem tabela no gabarito e não anula o sangramento dentro dele"
        if ".gabarito-body .table-wrap{" not in s else "sangramento anulado no gabarito")

# ─── G18 · o validador anuncia quantas conferências tem, e o número tem que
#      bater com a lista. A 1.3 dizia "Quatro conferências" e tinha cinco desde
#      que a onda 3-octies acrescentou uma. Ninguém conta ao acrescentar item.
titulo("G18 · O NÚMERO ANUNCIADO NO VALIDADOR BATE COM A LISTA")
NUM = {"uma": 1, "duas": 2, "três": 3, "quatro": 4, "cinco": 5,
       "seis": 6, "sete": 7, "oito": 8, "nove": 9, "dez": 10}
for f in AULAS:
    s = open(f, encoding="utf-8").read()
    m = re.search(r'class="checagem-lead">(.*?)</p>', s, re.S)
    itens = len(re.findall(r'<div class="checagem-o">', s))
    if not m:
        diz(False, "G18", f, "não tem checagem-lead"); continue
    lead = re.sub(r"<[^>]+>", "", m.group(1)).strip()
    prim = re.match(r"(\w+)", lead)
    dito = NUM.get(prim.group(1).lower()) if prim else None
    # lead que não abre com número não promete nada, e isso é legítimo
    if dito is None:   recado = f"{itens} itens, sem número anunciado"
    elif dito == itens: recado = f"{itens} itens"
    else:               recado = f"diz '{prim.group(1)}' e tem {itens}"
    diz(dito is None or dito == itens, "G18", f, recado)

# ─── G14 · a demonstração é passo a passo PARA O ALUNO, não roteiro de palco
titulo("G14 · DEMONSTRAÇÃO ESCRITA PARA O ALUNO, SEM DIREÇÃO DE CENA")
print("        (o site é do aluno · direção de palco vive fora dele · CLAUDE.md §9)")
# frases que só fazem sentido ditas pelo instrutor para a turma
PALCO = ["pergunte à sala", "pergunte a sala", "pergunte à turma", "plano b",
         "se der errado", "se a sala", "antes de a turma", "minutos de palco",
         "anote no quadro", "no quadro", "espere o silêncio", "a sala responde",
         "não responda você", "vire para a sala", "aponte na tela",
         "aponte isto", "diga em voz alta", "leia em voz alta", "a turma"]
for f in AULAS:
    dir_aula = os.path.dirname(f)
    demo = os.path.join(dir_aula, "demonstracao", "index.html")
    if not os.path.exists(demo):
        diz(False, "G14", f, "sem pasta demonstracao/"); continue
    bruto = open(demo, encoding="utf-8").read()
    limpo = re.sub(r"<[^>]+>", " ", bruto).lower()
    passos = len(re.findall(r'class="passo"', bruto))
    repare = len(re.findall(r'class="repare"', bruto))
    prompt = len(re.findall(r'class="pbox', bruto))
    # nenhum passo pode ser só instrução: todo passo tem que entregar alguma
    # coisa para o aluno levar (o que reparar, o que aquilo ensina, ou a lista)
    corpos = re.findall(r'<div class="passo-corpo">(.*?)\n    </div>', bruto, re.S)
    secos = [i + 1 for i, c in enumerate(corpos)
             if not re.search(r'class="(repare|ensina|lista)"', c)]
    ligado = 'href="./demonstracao/"' in open(f, encoding="utf-8").read()
    achados = sorted({t for t in PALCO if t in limpo})
    faltas = []
    if passos < 3:      faltas.append(f"só {passos} passos")
    if repare == 0:     faltas.append("nenhum bloco 'repare nisto'")
    if secos:           faltas.append(f"passo(s) só com instrução: {secos}")
    if prompt == 0:     faltas.append("nenhum prompt literal")
    if not ligado:      faltas.append("a aula não linka a demonstração")
    if achados:         faltas.append(f"DIREÇÃO DE CENA: {achados}")
    diz(not faltas, "G14", dir_aula,
        f"{passos} passos · {repare} repare · {prompt} prompts · 0 seco"
        if not faltas else str(faltas))

# ─────────── G14b · o G14 só olhava demonstracao/, e o defeito voltou pela porta
#             de fora: o card que aponta para ela continuava descrevendo roteiro.
#             gate que cobre parte da superfície deixa o erro voltar pelo resto.
titulo("G14b · NENHUMA PÁGINA DO ALUNO FALA COM O INSTRUTOR")
print("        (vale para TODO html do site, não só para a demonstração)")
# \b evita o falso positivo: "a sala" não pode casar dentro de "nesta sala"
CENA = [r"pergunte à (sala|turma)", r"pergunte a (sala|turma)", r"\bplano b\b",
        r"\ba sala\b (responde|dita|classific|precisa ver|conserta)",
        r"vire (para|pra) a sala", r"anote no quadro", r"espere o silêncio",
        r"não responda você", r"aponte (na tela|isto)", r"diga em voz alta",
        r"minutos de palco", r"se der errado", r"se a sala\b",
        # vocabulário do roteiro velho: a demonstração tem passos, não momentos
        r"\broteiro\b", r"momento \d+ d", r"os (três|quatro|cinco|seis) momentos",
        r"com o que apontar", r"executa ao vivo", r"travar na sala",
        # a plateia. "vocês" sozinho NÃO é defeito: "o dia a dia de vocês" e
        # "vocês já resolveram isso" tratam a turma como profissionais da
        # gráfica, e é dos melhores trechos do material. O defeito é a turma
        # como plateia assistindo alguém operar o teclado.
        r"na tela do rafael", r"pergunt\w* (para|pra) vocês",
        r"(na frente|diante) de vocês", r"mostr\w* (para|pra) vocês",
        r"vocês (veem|verão|vão ver|responderem|preenchem|acompanham)",
        r"o que eu quero", r"leio em voz alta", r"vamos contando"]
CENA_RE = re.compile("|".join(CENA), re.I)

# a lista de palavra não cobre PESSOA GRAMATICAL: "leia em voz alta" reprovava,
# "leio em voz alta" passava. Dentro do bloco que narra a demonstração, verbo em
# primeira pessoa do singular é o instrutor se descrevendo na página do aluno.
EU = re.compile(r"\b(Abro|Anexo|Peço|Corrijo|Mostro|Comparo|Mando|Rodo|Colo|Leio"
                r"|Explico|Pergunto|Escrevo|Faço|Jogo|Acrescento|Crio|Desligo"
                r"|Paro|Aponto|Digo|Chamo|Ligo|Uso|Gero|Repito|Volto)\b")
for f in HTML:
    bruto = open(f, encoding="utf-8").read()
    limpo = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", bruto))
    achados = sorted({m.group(0).strip() for m in CENA_RE.finditer(limpo)})
    # o bloco .demo descreve o que aparece na tela, nunca quem opera o teclado
    for demo in re.findall(r'<div class="demo-body">(.*?)</div>\s*</div>', bruto, re.S):
        texto = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", demo))
        achados += sorted({"1ª pessoa: " + m.group(0) for m in EU.finditer(texto)})
    achados = sorted(set(achados))
    diz(not achados, "G14b", f, f"FALA COM O INSTRUTOR: {achados}" if achados else "")

# ─────────── G15 · duração não vai para o material do aluno, em lugar nenhum
titulo("G15 · SEM DURAÇÃO EM NENHUMA PÁGINA DO ALUNO")
print("        (tempo é controle interno e vive fora do site · CLAUDE.md §9-ter)")
TEMPO = re.compile(r"\d+\s*(min\b|minutos?\b)", re.I)
SLOTS = [r'class="hero-kicker">([^<]*)<', r'class="destino-title">([^<]*)<',
         r'class="step-eyebrow">([^<]*)<', r'class="aula-dur">([^<]*)<',
         r'class="aula-title">([^<]*)<', r'class="kicker">([^<]*)<',
         r'class="passo-titulo">([^<]*)<']
for f in AULAS + ["m1/index.html"] + sorted(glob.glob("m1/a*/demonstracao/index.html")):
    s = open(f, encoding="utf-8").read()
    achados = [t for pat in SLOTS for t in re.findall(pat, s) if TEMPO.search(t)]
    diz(not achados, "G15", f, str(achados) if achados else "")

# ─── G20 · a MESMA skill aparece na aula 1.4 e na demonstração dela. Se as
#      duas descrições andarem separadas, a aula se contradiz a um clique de
#      distância. E o tamanho que a página afirma é recalculado do arquivo,
#      nunca digitado: eu tinha escrito "27 linhas" e são 19.
titulo("G20 · A SKILL DA 1.4 É A MESMA NA AULA E NA DEMONSTRAÇÃO")
import html as _html
A14  = "m1/a4-o-mapa/index.html"
D14  = "m1/a4-o-mapa/demonstracao/index.html"
if os.path.exists(A14) and os.path.exists(D14):
    dem = open(D14, encoding="utf-8").read()
    m = re.search(r'<div class="pconteudo" id="p1">(.*?)</div>', dem, re.S)
    diz(bool(m), "G20", "a demonstração tem o arquivo da skill",
        "" if m else "não achei o pconteudo p1")
    if m:
        skill  = _html.unescape(re.sub(r"<[^>]+>", "", m.group(1)))
        n_lin  = len(skill.split("\n"))
        n_pas  = len(re.findall(r"(?m)^\s*(\d+)\.", skill))
        nome   = re.search(r"(?m)^name:\s*(\S+)", skill)
        nome   = nome.group(1) if nome else ""
        print(f"        {'skill · linhas':28} {n_lin}")
        print(f"        {'skill · passos numerados':28} {n_pas}")
        print(f"        {'skill · nome':28} {nome}")
        aula = open(A14, encoding="utf-8").read()

        # 1. o nome é o mesmo nos dois lugares
        diz(nome and nome in aula, "G20", "a aula usa o mesmo nome de skill",
            f"a demo diz '{nome}'")

        # 2. presença não é conferência: cada "N linhas" dentro do rótulo do
        #    cabeçalho é conferido, não a existência do número em algum canto.
        cab = re.findall(r'class="anat-head">(.*?)</div>', aula, re.S)
        citados = [int(n) for b in cab for n in re.findall(r"(\d{1,4}) linhas", b)]
        diz(citados and all(c == n_lin for c in citados), "G20",
            "o tamanho no cabeçalho bate com o arquivo",
            f"a página diz {citados} e o arquivo tem {n_lin}"
            if citados != [n_lin] else f"{n_lin} linhas, conferido")

        # 3. os passos da aula são os mesmos passos da demonstração
        na = len(re.findall(r"<li>", re.search(
            r'<ol class="anat-passos">(.*?)</ol>', aula, re.S).group(1))) \
            if re.search(r'<ol class="anat-passos">', aula) else 0
        diz(na == n_pas, "G20", "a aula lista os mesmos passos da skill",
            f"aula={na} demo={n_pas}")

        # 4. as palavras-gatilho da descrição sobrevivem nos dois textos. É a
        #    descrição que faz a skill entrar, e é o que a aula inteira ensina.
        gatilhos = ["fecha essa cotação", "qual fornecedor eu escolho"]
        faltam = [g for g in gatilhos if g not in skill or g not in aula]
        diz(not faltam, "G20", "as palavras-gatilho estão nos dois lugares",
            str(faltam) if faltam else f"{len(gatilhos)} gatilhos")

# ─── G21 · o exemplo grande da 1.4 é um arquivo real fora do repo. Mesma
#      regra do G13c: a página afirma o tamanho, o gate recalcula, e se o
#      arquivo não existir nesta máquina o gate PULA alto, sem fingir que passou.
titulo("G21 · O TAMANHO DO EXEMPLO GRANDE SAI DO ARQUIVO")
ADV = "../insumos/exemplos/m1-aula-4/11-mentes-estrategicas/advisor-jeff-bezos.skill"
if os.path.exists(ADV):
    import zipfile as _zip
    z = _zip.ZipFile(ADV)
    dentro = [x for x in z.namelist() if x.endswith("SKILL.md")]
    diz(bool(dentro), "G21", "o .skill tem um SKILL.md dentro")
    if dentro:
        n_adv = len(z.read(dentro[0]).decode("utf-8").splitlines())
        print(f"        {'advisor · linhas':28} {n_adv}")
        aula = open(A14, encoding="utf-8").read()
        cab = re.findall(r'class="adv-head">(.*?)</div>', aula, re.S)
        citados = [int(n) for b in cab for n in re.findall(r"(\d{1,4}) linhas", b)]
        diz(citados == [n_adv], "G21", "o cabeçalho do exemplo bate com o arquivo",
            f"a página diz {citados} e o arquivo tem {n_adv}"
            if citados != [n_adv] else f"{n_adv} linhas, conferido")
        # Todo tamanho grande citado na 1.4 sai de um arquivo real, e são dois
        # arquivos diferentes: o exemplo grande daqui, e o campo de Instruções
        # da 1.3, que a 1.4 cita para dizer que tamanho não é critério. Sem
        # este segundo, o "128 linhas" da seção 01 é número solto: se aquele
        # arquivo encolher, a 1.4 mente e nada acusa.
        INST13 = "../insumos/exemplos/m1-aula-3/00-System_Instruction.md"
        validos = {n_adv}
        if os.path.exists(INST13):
            n_i13 = len(open(INST13, encoding="utf-8").read().splitlines())
            validos.add(n_i13)
            print(f"        {'instruções da 1.3 · linhas':28} {n_i13}")
            diz(f"{n_i13} linhas" in aula, "G21",
                "a 1.4 cita o tamanho real do exemplo da 1.3",
                f"precisa dizer '{n_i13} linhas'")
        else:
            print(f"  PULADO  {INST13} ausente, o número da 1.3 não foi conferido")
        outros = sorted({int(n) for n in re.findall(r"(\d{2,4}) linhas", aula)
                         if int(n) > 100 and int(n) not in validos})
        diz(not outros, "G21", "todo tamanho grande da 1.4 sai de um arquivo",
            f"{outros} não vêm de arquivo nenhum" if outros
            else f"conferidos contra {sorted(validos)}")
else:
    print(f"  PULADO  {ADV} não existe nesta máquina, o tamanho não foi conferido")

# ─── G22 · o nome do campo se copia da tela, nunca se traduz. Mesma lição do
#      G16, que nasceu quando "Contexto" virou "Conhecimento do projeto" em
#      seis lugares porque eu traduzi do help center em inglês.
titulo("G22 · A TELA DE SKILL É NOMEADA COMO A TELA NOMEIA")
print("        (é 'Habilidades' no claude.ai em pt-BR · verificado em 07/08/2026)")
if os.path.exists(A14):
    aula = open(A14, encoding="utf-8").read()
    tela = re.search(r'<div class="hab">(.*?)\n      </div>', aula, re.S)
    tela = tela.group(0) if tela else ""
    for alvo, onde in [("Habilidades", 'class="hab-tela">'),
                       ("por Você",    'class="hab-autor">')]:
        diz(f'{onde}{alvo}<' in tela, "G22", f"a tela diz '{alvo}'",
            "" if f'{onde}{alvo}<' in tela else f"não achei em {onde}")
    # a anatomia tem as três partes, com os nomes que a aula inteira usa
    partes = re.findall(r'class="anat-nome">([^<]*)<', aula)
    esperado = ["nome", "descrição", "o passo a passo"]
    diz(partes == esperado, "G22", "a anatomia mostra as três partes na ordem",
        f"achei {partes}" if partes != esperado else "nome · descrição · passo a passo")
    # e a página do aluno não pode chamar a tela pelo nome em inglês
    ingles = [t for t in ["Skills tab", "aba Skills", "menu Skills"] if t in aula]
    diz(not ingles, "G22", "não usa o nome em inglês da tela",
        str(ingles) if ingles else "")

# ─── G23 · "volte ao passo N" aponta para passo que existe. Nasceu quando o
#      passo 5 da 1.4 virou três passos e o validador continuou mandando voltar
#      ao 5, que agora é outro. Referência a número de passo dessincroniza calada.
#
#      ⚠️ O que este gate NÃO faz: ele confere a FAIXA, não o alvo. Se a lista
#      cresce e a referência antiga continua dentro da faixa, ele passa. Para o
#      alvo certo não existe checagem mecânica, é leitura.
titulo("G23 · REFERÊNCIA A PASSO APONTA PARA PASSO QUE EXISTE")
for f in AULAS:
    s = open(f, encoding="utf-8").read()
    # duas listas de passos convivem na mesma página: a do exercício e a da
    # demonstração. Uma referência vale se couber em alguma das duas.
    n_ex = len(re.findall(r'<li>\s*<div class="passo-title">', s))
    n_demo = max((len(re.findall(r"<li>", bloco)) for bloco in
                  re.findall(r'<ol class="demo-passos">(.*?)</ol>', s, re.S)), default=0)
    teto = max(n_ex, n_demo)
    citados = [int(n) for n in re.findall(r"\bpasso (\d+)\b", s)]
    fora = sorted({n for n in citados if n < 1 or n > teto})
    diz(not fora, "G23", f,
        f"aponta para {fora}, e a maior lista tem {teto}" if fora
        else f"{len(citados)} referências · exercício={n_ex} demo={n_demo}")

# ─── G24 · as oito regras da Gráfica Aurora existem em DOIS lugares na 1.4:
#      no quadro preenchível (array do JS) e no gabarito (tabela HTML). Se as
#      duas listas andarem separadas, a pessoa responde uma coisa e confere
#      outra. Mesma família do G20: cópia da mesma verdade em dois arquivos.
titulo("G24 · O QUADRO E O GABARITO CLASSIFICAM AS MESMAS OITO REGRAS")
if os.path.exists(A14):
    aula = open(A14, encoding="utf-8").read()
    m = re.search(r"var FIXAS = \[(.*?)\];", aula, re.S)
    diz(bool(m), "G24", "o quadro declara as oito regras", "" if m else "sem array FIXAS")
    if m:
        quadro = re.findall(r"'([^']+)'", m.group(1))
        gab = re.search(r'<details class="gabarito".*?</details>', aula, re.S)
        linhas = re.findall(r"<tr>\s*<td>([^<]+)</td>", gab.group(0), re.S) if gab else []
        gabarito = [re.sub(r"\s+", " ", t).strip() for t in linhas]
        print(f"        {'quadro':28} {len(quadro)} regras")
        print(f"        {'gabarito':28} {len(gabarito)} regras")
        diz(len(quadro) == 8, "G24", "o quadro tem exatamente oito", f"{len(quadro)}")
        diz(quadro == gabarito, "G24", "as duas listas batem, na mesma ordem",
            f"diferem em {[i + 1 for i, (a, b) in enumerate(zip(quadro, gabarito)) if a != b]}"
            if quadro != gabarito else "8 de 8, mesma ordem")
        # e o texto do exercício não pode prometer outro número. A checagem
        # exige ACHAR menção: "0 menções" passando é passar no vazio, e foi o
        # que a primeira versão desta linha fazia.
        POR_EXTENSO = {6: "seis", 7: "sete", 8: "oito", 9: "nove", 10: "dez"}
        certo = POR_EXTENSO[len(quadro)]
        n_txt = re.findall(r"\b(seis|sete|oito|nove|dez)\s+(?:regras\s+)?da Gráfica Aurora", aula)
        diz(len(n_txt) >= 2 and all(t == certo for t in n_txt), "G24",
            f"o texto da aula promete '{certo}' em todo lugar",
            f"achei {n_txt}, e são {len(quadro)} regras" if n_txt.count(certo) != len(n_txt)
            else f"{len(n_txt)} menções, todas '{certo}'"
            if len(n_txt) >= 2 else "menos de 2 menções, a checagem passaria no vazio")

# ─────────── o fecho que faltava
# A lista `falhas` existia desde a onda 1 e NADA a lia: o script imprimia
# FALHA e saia com codigo 0. O README prometia o contrario havia quatro
# ondas. Gate que reprova sem falhar a execucao nao e gate, e aviso.
print()
print("═" * 74)
if falhas:
    print(f"  {len(falhas)} FALHA(S):")
    for f in falhas:
        print(f"    · {f}")
    print("═" * 74)
    sys.exit(1)
print("  TUDO OK · nenhuma falha")
print("═" * 74)
